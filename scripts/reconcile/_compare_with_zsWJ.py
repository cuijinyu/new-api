"""
对比知书万卷账单 vs 我方降档后金额（3月19-24日, 渠道54）
使用 gen_bill.py 的降档逻辑重算 expected_usd
"""
import sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
os.chdir(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
import sqlite3
import pandas as pd
import numpy as np
from collections import defaultdict
from datetime import datetime, timezone, timedelta

from gen_bill import (
    parse_other_batch, assign_prices, compute_costs,
    FLAT_TIER_MODELS, PRICING, TIER_DF, FLAT_DF, QUOTA_PER_USD,
    DB_PATH
)

BJ = timezone(timedelta(hours=8))
BILL_PATH = "reconcile/aws5101消费记录（2026-03-19 00-00-00 - 2026-03-25 00-00-00）.xlsx"

# 3月13日 00:00 北京时间 = UTC timestamp
FLAT_TIER_SINCE_TS = int(datetime(2026, 3, 13, tzinfo=BJ).timestamp())

def ts_to_date(ts):
    return datetime.fromtimestamp(ts, tz=BJ).strftime("%m-%d")

# ── 1. 解析知书万卷账单 ──
print("Loading their bill...")
wb = openpyxl.load_workbook(BILL_PATH, read_only=True, data_only=True)
ws = wb[wb.sheetnames[0]]

their_model = defaultdict(lambda: {"count": 0, "quota": 0})
their_date_model = defaultdict(lambda: defaultdict(lambda: {"count": 0, "quota": 0}))
ts_min, ts_max = float("inf"), 0

for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0:
        continue
    rid, user_id, created_at, typ = row[:4]
    model_name, quota = row[7], row[8]
    if typ != 2:
        continue
    q = quota or 0
    their_model[model_name]["count"] += 1
    their_model[model_name]["quota"] += q
    date = ts_to_date(created_at)
    their_date_model[date][model_name]["count"] += 1
    their_date_model[date][model_name]["quota"] += q
    ts_min = min(ts_min, created_at)
    ts_max = max(ts_max, created_at)
wb.close()

# ── 2. 从数据库加载同时间段渠道54数据，用降档逻辑重算 ──
print("Loading our data from DB...")
conn = sqlite3.connect(DB_PATH)
raw = pd.read_sql_query(
    """SELECT id, model_name AS model, quota, prompt_tokens, completion_tokens,
              other, token_name, created_at,
              datetime(created_at,'unixepoch','+8 hours') AS cst
       FROM logs
       WHERE channel_id=54 AND type=2
         AND created_at>=? AND created_at<=?
       ORDER BY created_at""",
    conn, params=(ts_min, ts_max)
)
conn.close()

print(f"Our rows: {len(raw):,}")

other_df = parse_other_batch(raw["other"])
df = pd.concat([raw.drop(columns=["other"]), other_df], axis=1)

# 降档：3月13日起 opus-4-6/sonnet-4-6 用低档价
df = assign_prices(df, flat_tier=True, flat_tier_since_ts=FLAT_TIER_SINCE_TS)
df = df.dropna(subset=["ip"])
df = compute_costs(df, flat_tier=True, flat_tier_since_ts=FLAT_TIER_SINCE_TS)

# 加日期列
df["date"] = df["created_at"].apply(ts_to_date)

# ── 3. 按模型汇总对比 ──
our_model = df.groupby("model").agg(
    count=("id", "count"),
    billed_usd=("billed_usd", "sum"),
    expected_usd=("expected_usd", "sum"),
).reset_index()

all_models = sorted(set(list(their_model.keys()) + list(our_model["model"])))

print("\n" + "=" * 140)
print("知书万卷账单 vs 我方降档后金额 对比 (3/19-3/24, ch=54)")
print("=" * 140)
print(f"{'模型':<35} {'知书次数':>10} {'我方次数':>10} {'知书$(quota)':>14} {'我方billed$':>14} {'我方降档$':>14} {'差额(降档)':>12} {'差%':>8}")
print("-" * 140)

sum_t = sum_ob = sum_oe = 0
for m in all_models:
    t = their_model.get(m, {"count": 0, "quota": 0})
    t_usd = t["quota"] / QUOTA_PER_USD
    o_row = our_model[our_model["model"] == m]
    if len(o_row):
        o_cnt = int(o_row["count"].iloc[0])
        o_billed = float(o_row["billed_usd"].iloc[0])
        o_expected = float(o_row["expected_usd"].iloc[0])
    else:
        o_cnt = o_billed = o_expected = 0

    diff = o_expected - t_usd
    pct = diff / t_usd * 100 if t_usd else 0
    sum_t += t_usd
    sum_ob += o_billed
    sum_oe += o_expected
    print(f"  {m:<33} {t['count']:>10,} {o_cnt:>10,} {t_usd:>14,.2f} {o_billed:>14,.2f} {o_expected:>14,.2f} {diff:>+12,.2f} {pct:>+7.1f}%")

print("-" * 140)
total_diff = sum_oe - sum_t
total_pct = total_diff / sum_t * 100 if sum_t else 0
print(f"  {'TOTAL':<33} {'':>10} {'':>10} {sum_t:>14,.2f} {sum_ob:>14,.2f} {sum_oe:>14,.2f} {total_diff:>+12,.2f} {total_pct:>+7.1f}%")

# ── 4. 按日期+模型对比（重点模型） ──
our_date_model = df.groupby(["date", "model"]).agg(
    count=("id", "count"),
    billed_usd=("billed_usd", "sum"),
    expected_usd=("expected_usd", "sum"),
).reset_index()

all_dates = sorted(set(list(their_date_model.keys()) + list(df["date"].unique())))

for focus in ["claude-opus-4-6", "claude-sonnet-4-6"]:
    print(f"\n{'='*120}")
    print(f"按日期对比: {focus}")
    print(f"{'='*120}")
    print(f"{'日期':<10} {'知书次数':>10} {'我方次数':>10} {'知书$':>14} {'我方billed$':>14} {'我方降档$':>14} {'差额(降档)':>12} {'差%':>8}")
    print("-" * 100)
    for d in all_dates:
        t = their_date_model[d].get(focus, {"count": 0, "quota": 0})
        t_usd = t["quota"] / QUOTA_PER_USD
        o_row = our_date_model[(our_date_model["date"] == d) & (our_date_model["model"] == focus)]
        if len(o_row):
            o_cnt = int(o_row["count"].iloc[0])
            o_billed = float(o_row["billed_usd"].iloc[0])
            o_expected = float(o_row["expected_usd"].iloc[0])
        else:
            o_cnt = o_billed = o_expected = 0
        diff = o_expected - t_usd
        pct = diff / t_usd * 100 if t_usd else 0
        print(f"  {d:<8} {t['count']:>10,} {o_cnt:>10,} {t_usd:>14,.2f} {o_billed:>14,.2f} {o_expected:>14,.2f} {diff:>+12,.2f} {pct:>+7.1f}%")

# ── 5. 按日期全模型合计 ──
print(f"\n{'='*120}")
print("按日期对比: 所有模型合计")
print(f"{'='*120}")
print(f"{'日期':<10} {'知书次数':>10} {'我方次数':>10} {'知书$':>14} {'我方billed$':>14} {'我方降档$':>14} {'差额(降档)':>12} {'差%':>8}")
print("-" * 100)
for d in all_dates:
    t_cnt = sum(v["count"] for v in their_date_model[d].values())
    t_usd = sum(v["quota"] for v in their_date_model[d].values()) / QUOTA_PER_USD
    o_day = our_date_model[our_date_model["date"] == d]
    o_cnt = int(o_day["count"].sum()) if len(o_day) else 0
    o_billed = float(o_day["billed_usd"].sum()) if len(o_day) else 0
    o_expected = float(o_day["expected_usd"].sum()) if len(o_day) else 0
    diff = o_expected - t_usd
    pct = diff / t_usd * 100 if t_usd else 0
    print(f"  {d:<8} {t_cnt:>10,} {o_cnt:>10,} {t_usd:>14,.2f} {o_billed:>14,.2f} {o_expected:>14,.2f} {diff:>+12,.2f} {pct:>+7.1f}%")
