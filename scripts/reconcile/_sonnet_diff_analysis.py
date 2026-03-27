"""
深入分析 sonnet-4-6 在知书万卷账单 vs 我方数据库的差异来源
按日期逐条对比 quota 分布、次数差异、cache 计费差异
"""
import sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
os.chdir(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
import sqlite3
import json
import pandas as pd
import numpy as np
from collections import defaultdict
from datetime import datetime, timezone, timedelta

from gen_bill import (
    parse_other_batch, assign_prices, compute_costs,
    FLAT_TIER_MODELS, QUOTA_PER_USD, DB_PATH
)

BJ = timezone(timedelta(hours=8))
BILL_PATH = "reconcile/aws5101消费记录（2026-03-19 00-00-00 - 2026-03-25 00-00-00）.xlsx"
FLAT_TIER_SINCE_TS = int(datetime(2026, 3, 13, tzinfo=BJ).timestamp())
MODEL = "claude-sonnet-4-6"

def ts_to_date(ts):
    return datetime.fromtimestamp(ts, tz=BJ).strftime("%m-%d")

# ── 1. 解析知书万卷 sonnet-4-6 ──
print(f"Loading their bill for {MODEL}...")
wb = openpyxl.load_workbook(BILL_PATH, read_only=True, data_only=True)
ws = wb[wb.sheetnames[0]]

their_rows = []
ts_min, ts_max = float("inf"), 0
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0:
        headers = list(row)
        continue
    model_name = row[7]
    if model_name != MODEL:
        continue
    rid, user_id, created_at, typ = row[0], row[1], row[2], row[3]
    quota, prompt_tokens, completion_tokens = row[8], row[9], row[10]
    other_str = row[-1] if len(row) >= 18 else None
    if typ != 2:
        continue
    their_rows.append({
        "id": rid, "created_at": created_at,
        "quota": quota or 0,
        "prompt_tokens": prompt_tokens or 0,
        "completion_tokens": completion_tokens or 0,
        "other": other_str or "",
        "date": ts_to_date(created_at),
    })
    ts_min = min(ts_min, created_at)
    ts_max = max(ts_max, created_at)
wb.close()

their_df = pd.DataFrame(their_rows)
print(f"Their {MODEL}: {len(their_df):,} rows")

# ── 2. 我方数据 ──
print(f"Loading our data for {MODEL}...")
conn = sqlite3.connect(DB_PATH)
our_raw = pd.read_sql_query(
    """SELECT id, model_name AS model, quota, prompt_tokens, completion_tokens,
              other, token_name, created_at,
              datetime(created_at,'unixepoch','+8 hours') AS cst
       FROM logs
       WHERE channel_id=54 AND type=2
         AND model_name=?
         AND created_at>=? AND created_at<=?
       ORDER BY created_at""",
    conn, params=(MODEL, ts_min, ts_max)
)
conn.close()
print(f"Our {MODEL}: {len(our_raw):,} rows")

# ── 3. 按日期对比 quota 总额 ──
our_raw["date"] = our_raw["created_at"].apply(ts_to_date)

their_daily = their_df.groupby("date").agg(
    t_count=("id", "count"),
    t_quota=("quota", "sum"),
    t_prompt=("prompt_tokens", "sum"),
    t_completion=("completion_tokens", "sum"),
).reset_index()

our_daily = our_raw.groupby("date").agg(
    o_count=("id", "count"),
    o_quota=("quota", "sum"),
    o_prompt=("prompt_tokens", "sum"),
    o_completion=("completion_tokens", "sum"),
).reset_index()

daily = their_daily.merge(our_daily, on="date", how="outer").fillna(0)
daily["diff_count"] = daily["o_count"] - daily["t_count"]
daily["diff_quota"] = daily["o_quota"] - daily["t_quota"]
daily["t_usd"] = daily["t_quota"] / QUOTA_PER_USD
daily["o_usd"] = daily["o_quota"] / QUOTA_PER_USD
daily["diff_usd"] = daily["o_usd"] - daily["t_usd"]

print(f"\n{'='*120}")
print(f"sonnet-4-6 按日期对比 (quota = 系统 billed)")
print(f"{'='*120}")
print(f"{'date':<8} {'t_cnt':>8} {'o_cnt':>8} {'d_cnt':>7} {'t_prompt':>14} {'o_prompt':>14} {'t_comp':>12} {'o_comp':>12} {'t_$':>12} {'o_$':>12} {'diff_$':>10}")
print("-" * 120)
for _, r in daily.iterrows():
    print(f"  {r['date']:<6} {int(r['t_count']):>8,} {int(r['o_count']):>8,} {int(r['diff_count']):>+7,} "
          f"{int(r['t_prompt']):>14,} {int(r['o_prompt']):>14,} "
          f"{int(r['t_completion']):>12,} {int(r['o_completion']):>12,} "
          f"{r['t_usd']:>12,.2f} {r['o_usd']:>12,.2f} {r['diff_usd']:>+10,.2f}")

# ── 4. 分析 prompt_tokens 分布差异 ──
print(f"\n{'='*120}")
print("sonnet-4-6 prompt_tokens 分布对比")
print(f"{'='*120}")

bins = [0, 10000, 50000, 100000, 200000, 500000, float("inf")]
labels = ["0-10k", "10k-50k", "50k-100k", "100k-200k", "200k-500k", "500k+"]

their_df["pt_bin"] = pd.cut(their_df["prompt_tokens"], bins=bins, labels=labels, right=False)
our_raw["pt_bin"] = pd.cut(our_raw["prompt_tokens"], bins=bins, labels=labels, right=False)

t_dist = their_df.groupby("pt_bin", observed=False).agg(cnt=("id","count"), quota=("quota","sum")).reset_index()
o_dist = our_raw.groupby("pt_bin", observed=False).agg(cnt=("id","count"), quota=("quota","sum")).reset_index()

print(f"{'bin':<15} {'t_cnt':>10} {'o_cnt':>10} {'d_cnt':>8} {'t_$':>12} {'o_$':>12} {'diff_$':>10}")
print("-" * 80)
for i in range(len(t_dist)):
    t_c, t_q = int(t_dist.iloc[i]["cnt"]), t_dist.iloc[i]["quota"]/QUOTA_PER_USD
    o_c, o_q = int(o_dist.iloc[i]["cnt"]), o_dist.iloc[i]["quota"]/QUOTA_PER_USD
    print(f"  {labels[i]:<13} {t_c:>10,} {o_c:>10,} {o_c-t_c:>+8,} {t_q:>12,.2f} {o_q:>12,.2f} {o_q-t_q:>+10,.2f}")

# ── 5. 用降档逻辑重算我方 expected_usd，和知书万卷 quota 对比 ──
print(f"\n{'='*120}")
print("sonnet-4-6 降档后 expected_usd vs 知书万卷 quota")
print(f"{'='*120}")

other_df = parse_other_batch(our_raw["other"])
df = pd.concat([our_raw.drop(columns=["other"]), other_df], axis=1)
df = assign_prices(df, flat_tier=True, flat_tier_since_ts=FLAT_TIER_SINCE_TS)
df = df.dropna(subset=["ip"])
df = compute_costs(df, flat_tier=True, flat_tier_since_ts=FLAT_TIER_SINCE_TS)

our_daily2 = df.groupby("date").agg(
    o_count=("id", "count"),
    o_billed=("billed_usd", "sum"),
    o_expected=("expected_usd", "sum"),
    o_cost_input=("cost_input", "sum"),
    o_cost_output=("cost_output", "sum"),
    o_cost_cache_hit=("cost_cache_hit", "sum"),
    o_cost_cw_5m=("cost_cw_5m", "sum"),
    o_cost_cw_1h=("cost_cw_1h", "sum"),
).reset_index()

daily2 = their_daily.merge(our_daily2, on="date", how="outer").fillna(0)
daily2["t_usd"] = daily2["t_quota"] / QUOTA_PER_USD

print(f"{'date':<8} {'t_$(quota)':>12} {'o_billed$':>12} {'o_expected$':>12} {'t-o_exp':>10} {'input$':>10} {'output$':>10} {'ch_hit$':>10} {'cw_5m$':>10} {'cw_1h$':>10}")
print("-" * 120)
for _, r in daily2.iterrows():
    diff = r["t_usd"] - r["o_expected"]
    print(f"  {r['date']:<6} {r['t_usd']:>12,.2f} {r['o_billed']:>12,.2f} {r['o_expected']:>12,.2f} "
          f"{diff:>+10,.2f} {r['o_cost_input']:>10,.2f} {r['o_cost_output']:>10,.2f} "
          f"{r['o_cost_cache_hit']:>10,.2f} {r['o_cost_cw_5m']:>10,.2f} {r['o_cost_cw_1h']:>10,.2f}")

# ── 6. 对比 cache 相关 tokens ──
print(f"\n{'='*120}")
print("sonnet-4-6 cache tokens 对比")
print(f"{'='*120}")

# 知书万卷的 other 字段解析
def parse_their_cache(other_str):
    try:
        o = json.loads(other_str) if other_str else {}
    except:
        o = {}
    return {
        "cache_tokens": o.get("cache_tokens", 0) or 0,
        "cache_creation_tokens": o.get("cache_creation_tokens", 0) or 0,
    }

t_cache = their_df["other"].apply(lambda x: pd.Series(parse_their_cache(x)))
their_df["t_ch"] = t_cache["cache_tokens"]
their_df["t_cw"] = t_cache["cache_creation_tokens"]

t_cache_daily = their_df.groupby("date").agg(
    t_ch=("t_ch", "sum"), t_cw=("t_cw", "sum"),
).reset_index()

o_cache_daily = df.groupby("date").agg(
    o_ch=("ch", "sum"), o_cw_5m=("cw_5m_total", "sum"), o_cw_1h=("cw_1h", "sum"),
).reset_index()

cache_cmp = t_cache_daily.merge(o_cache_daily, on="date", how="outer").fillna(0)
print(f"{'date':<8} {'t_cache_hit':>14} {'o_cache_hit':>14} {'diff':>12} {'t_cache_write':>14} {'o_cw_5m':>14} {'o_cw_1h':>14}")
print("-" * 100)
for _, r in cache_cmp.iterrows():
    print(f"  {r['date']:<6} {int(r['t_ch']):>14,} {int(r['o_ch']):>14,} {int(r['o_ch']-r['t_ch']):>+12,} "
          f"{int(r['t_cw']):>14,} {int(r['o_cw_5m']):>14,} {int(r['o_cw_1h']):>14,}")

# ── 7. 知书万卷 sonnet 的 model_ratio / model_price 分析 ──
print(f"\n{'='*120}")
print("知书万卷 sonnet-4-6 的 model_ratio 分布")
print(f"{'='*120}")

def extract_ratio(other_str):
    try:
        o = json.loads(other_str) if other_str else {}
    except:
        o = {}
    return o.get("model_ratio", None), o.get("completion_ratio", None), o.get("cache_ratio", None)

ratios = their_df["other"].apply(lambda x: pd.Series(extract_ratio(x), index=["model_ratio","comp_ratio","cache_ratio"]))
for col in ["model_ratio", "comp_ratio", "cache_ratio"]:
    vc = ratios[col].value_counts().head(10)
    print(f"\n{col}:")
    for v, c in vc.items():
        print(f"  {v}: {c:,} rows")

# ── 8. 用知书万卷的 ratio 重算他们的 quota ──
print(f"\n{'='*120}")
print("知书万卷 sonnet-4-6 计费逻辑验证")
print(f"{'='*120}")
print("验证: quota = (prompt * model_ratio + completion * model_ratio * comp_ratio + cache_hit * model_ratio * cache_ratio + cache_write * model_ratio * 1.25) * 2")
print("(基础单位: 1 token = 1/500000 USD 的 quota)")

sample = their_df.head(20).copy()
for idx, row in sample.iterrows():
    try:
        o = json.loads(row["other"]) if row["other"] else {}
    except:
        o = {}
    mr = o.get("model_ratio", 1)
    cr = o.get("completion_ratio", 1)
    chr_ = o.get("cache_ratio", 0)
    cwr = o.get("cache_creation_ratio", 1)
    ch = o.get("cache_tokens", 0) or 0
    cw = o.get("cache_creation_tokens", 0) or 0
    pt = row["prompt_tokens"]
    ct = row["completion_tokens"]
    net_input = max(pt - ch - cw, 0)

    # 知书万卷的 quota 计算: base_ratio=2 (group_ratio=1, 但 quota 基数是 2/token?)
    # 尝试几种公式
    calc1 = (net_input * mr + ct * mr * cr + ch * mr * chr_ + cw * mr * cwr) * 2
    calc2 = (pt * mr + ct * mr * cr) * 2
    calc3 = net_input * mr + ct * mr * cr + ch * mr * chr_ + cw * mr * cwr

    actual = row["quota"]
    match1 = "OK" if abs(calc1 - actual) < 2 else ""
    match2 = "OK" if abs(calc2 - actual) < 2 else ""
    match3 = "OK" if abs(calc3 - actual) < 2 else ""

    print(f"  id={row['id']} pt={pt} ct={ct} ch={ch} cw={cw} mr={mr} cr={cr} chr={chr_} cwr={cwr} "
          f"actual={actual} calc1={calc1:.0f}{' '+match1 if match1 else ''} "
          f"calc2={calc2:.0f}{' '+match2 if match2 else ''} "
          f"calc3={calc3:.0f}{' '+match3 if match3 else ''}")
