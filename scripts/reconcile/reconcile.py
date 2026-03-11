#!/usr/bin/env python3
"""
每日对账脚本：从 S3 原始日志中提取 usage 信息，按模型价格独立计算费用。

支持：
  - 分段计费 (tiered_pricing)
  - Claude 200K 自动倍率 (input x2, output x1.5)
  - 缓存写入 5m / 1h TTL 区分
  - Claude / OpenAI Web Search 工具调用计费

用法:
    python reconcile.py                          # 对账昨天
    python reconcile.py --date 2026-03-10        # 对账指定日期
    python reconcile.py --date 2026-03-10 --output report.csv
    python reconcile.py --date 2026-03-10 --bill bill.xlsx          # 导出账单
    python reconcile.py --date 2026-03-10 --bill bill.xlsx --user-id 1  # 指定用户账单

环境变量 (或通过命令行参数):
    S3_BUCKET, S3_REGION, S3_PREFIX, S3_ENDPOINT,
    AWS_ACCESS_KEY_ID, AWS_SECRET_ACCESS_KEY
"""

import argparse
import csv
import gzip
import json
import os
import re
import sys
import time
from collections import defaultdict, Counter
from concurrent.futures import ProcessPoolExecutor, ThreadPoolExecutor, as_completed
from datetime import datetime, timedelta

import boto3
from tabulate import tabulate

# Claude 200K 阈值与倍率，对齐 ratio_setting/model_ratio.go
CLAUDE_200K_THRESHOLD = 200_000
CLAUDE_200K_INPUT_MULT = 2.0
CLAUDE_200K_OUTPUT_MULT = 1.5

CLAUDE_MODEL_RE = re.compile(r"claude", re.IGNORECASE)


def parse_args():
    p = argparse.ArgumentParser(description="S3 原始日志费用计算")
    p.add_argument("--date", type=str, default=None,
                   help="对账日期 (YYYY-MM-DD)，默认昨天")
    p.add_argument("--date-range", type=str, nargs=2, metavar=("START", "END"),
                   help="日期范围 (YYYY-MM-DD YYYY-MM-DD)")
    p.add_argument("--bucket", type=str,
                   default=os.getenv("RAW_LOG_S3_BUCKET", os.getenv("S3_BUCKET", "")))
    p.add_argument("--region", type=str,
                   default=os.getenv("RAW_LOG_S3_REGION", os.getenv("S3_REGION", "us-east-1")))
    p.add_argument("--prefix", type=str,
                   default=os.getenv("RAW_LOG_S3_PREFIX", os.getenv("S3_PREFIX", "llm-raw-logs")))
    p.add_argument("--endpoint", type=str,
                   default=os.getenv("RAW_LOG_S3_ENDPOINT", os.getenv("S3_ENDPOINT", "")))
    p.add_argument("--pricing", type=str, default="pricing.json",
                   help="模型价格配置文件路径")
    p.add_argument("--output", type=str, default=None,
                   help="输出 CSV 文件路径")
    p.add_argument("--group-by", type=str, default="model",
                   choices=["model", "channel", "user", "hour"],
                   help="汇总维度")
    p.add_argument("--user-id", type=int, nargs="+", default=None,
                   help="按用户 ID 过滤（可指定多个）")
    p.add_argument("--model", type=str, nargs="+", default=None,
                   help="按模型名过滤（可指定多个）")
    p.add_argument("--channel-id", type=int, nargs="+", default=None,
                   help="按渠道 ID 过滤（可指定多个）")
    p.add_argument("--bill", type=str, default=None,
                   help="导出 Excel 账单文件路径 (.xlsx)")
    p.add_argument("--bill-title", type=str, default="API 使用账单",
                   help="账单标题")
    p.add_argument("--bill-currency", type=str, default="USD",
                   choices=["USD", "CNY"],
                   help="账单币种 (USD 或 CNY)")
    p.add_argument("--exchange-rate", type=float, default=7.3,
                   help="USD 转 CNY 汇率，默认 7.3")
    p.add_argument("--workers", type=int, default=10,
                   help="并发下载线程数，默认 10")
    p.add_argument("--processes", type=int, default=0,
                   help="并发进程数，默认 0（禁用；建议大量小文件时设置为 CPU 核数）")
    p.add_argument("--cache-dir", type=str, default=".cache",
                   help="本地缓存目录，缓存已下载的 S3 文件（默认 .cache）")
    p.add_argument("--no-cache", action="store_true",
                   help="禁用本地缓存，强制从 S3 下载")
    p.add_argument("--time-from", type=str, default=None,
                   help="起始时间过滤，格式 HH:MM:SS 或 YYYY-MM-DDTHH:MM:SS")
    p.add_argument("--time-to", type=str, default=None,
                   help="截止时间过滤，格式 HH:MM:SS 或 YYYY-MM-DDTHH:MM:SS")
    p.add_argument("--verbose", action="store_true")
    return p.parse_args()


def load_pricing(filepath):
    with open(filepath, "r", encoding="utf-8") as f:
        data = json.load(f)
    return data


def get_s3_client(region, endpoint):
    kwargs = {"region_name": region}
    if endpoint:
        kwargs["endpoint_url"] = endpoint
    ak = os.getenv("RAW_LOG_S3_ACCESS_KEY_ID", os.getenv("AWS_ACCESS_KEY_ID"))
    sk = os.getenv("RAW_LOG_S3_SECRET_ACCESS_KEY", os.getenv("AWS_SECRET_ACCESS_KEY"))
    if ak and sk:
        kwargs["aws_access_key_id"] = ak
        kwargs["aws_secret_access_key"] = sk
    return boto3.client("s3", **kwargs)


def list_s3_objects(s3, bucket, prefix):
    objects = []
    token = None
    while True:
        kw = {"Bucket": bucket, "Prefix": prefix, "MaxKeys": 1000}
        if token:
            kw["ContinuationToken"] = token
        resp = s3.list_objects_v2(**kw)
        for obj in resp.get("Contents", []):
            objects.append(obj["Key"])
        if not resp.get("IsTruncated"):
            break
        token = resp["NextContinuationToken"]
    return objects


def _parse_gzip_data(raw):
    """解析 gzip 压缩的 JSONL 数据"""
    data = gzip.decompress(raw)
    records = []
    for line in data.decode("utf-8", errors="replace").strip().split("\n"):
        if not line.strip():
            continue
        try:
            records.append(json.loads(line))
        except json.JSONDecodeError:
            continue
    return records


def _get_cache_path(cache_dir, key):
    """根据 S3 key 生成本地缓存文件路径"""
    return os.path.join(cache_dir, key)


def download_and_parse(s3, bucket, key, cache_dir=None):
    if cache_dir:
        cache_path = _get_cache_path(cache_dir, key)
        if os.path.exists(cache_path):
            with open(cache_path, "rb") as f:
                return _parse_gzip_data(f.read())

    resp = s3.get_object(Bucket=bucket, Key=key)
    raw = resp["Body"].read()

    if cache_dir:
        cache_path = _get_cache_path(cache_dir, key)
        os.makedirs(os.path.dirname(cache_path), exist_ok=True)
        with open(cache_path, "wb") as f:
            f.write(raw)

    return _parse_gzip_data(raw)


# ---------------------------------------------------------------------------
# Usage 提取
# ---------------------------------------------------------------------------

def classify_error_response(body, status_code):
    """从 response_body 中识别上游错误类型，返回可读的分类字符串"""
    error_type = ""
    error_msg = ""

    # SSE 格式的 error event（data: 后可能跨多行）
    lines = body.split("\n")
    i = 0
    while i < len(lines):
        line = lines[i].strip()
        if line.startswith("data:"):
            payload = line[len("data:"):].strip()
            # 收集可能的多行 JSON（后续非 event:/data:/空行 的行属于同一 data 块）
            j = i + 1
            while j < len(lines):
                next_line = lines[j].strip()
                if not next_line or next_line.startswith("event:") or next_line.startswith("data:"):
                    break
                payload += next_line
                j += 1
            i = j
            try:
                d = json.loads(payload)
                err = d.get("error", {})
                if isinstance(err, dict):
                    error_type = err.get("type", "")
                    error_msg = err.get("message", "")
                    if error_type:
                        break
            except (json.JSONDecodeError, AttributeError):
                continue
        else:
            i += 1

    # 非 SSE 的 JSON error
    if not error_type:
        try:
            obj = json.loads(body)
            err = obj.get("error", {})
            if isinstance(err, dict):
                error_type = err.get("type", err.get("code", ""))
                error_msg = err.get("message", "")
            elif isinstance(err, str):
                error_type = err
        except (json.JSONDecodeError, AttributeError):
            pass

    if error_type:
        short_msg = _shorten_error_msg(error_msg)
        label = f"{error_type}"
        if short_msg:
            label += f": {short_msg}"
        return label

    if status_code and status_code >= 400:
        return f"http_{status_code}"

    return "unknown_error"


def _shorten_error_msg(msg):
    """将冗长的错误消息缩短为可读的分类标签"""
    if not msg:
        return ""
    if "tool_use" in msg and "tool_result" in msg:
        return "tool_use缺少tool_result"
    if "credit" in msg.lower() or "balance" in msg.lower() or "quota" in msg.lower():
        return "额度不足"
    if "rate_limit" in msg.lower() or "rate limit" in msg.lower():
        return "速率限制"
    if "context_length" in msg.lower() or "too many tokens" in msg.lower():
        return "上下文超长"
    if "overloaded" in msg.lower():
        return "服务过载"
    if "timeout" in msg.lower():
        return "超时"
    if len(msg) > 60:
        return msg[:57] + "..."
    return msg


def extract_usage(record):
    """
    从 response_body 中提取上游返回的 token 用量。
    成功返回 dict，失败返回 (None, reason_str)。
    """
    body = record.get("response_body", "")
    status_code = record.get("status_code", 0)

    if not body:
        if status_code and status_code >= 400:
            return None, f"http_{status_code}_empty_body"
        return None, "empty_response_body"

    # 非流式：完整 JSON
    try:
        obj = json.loads(body)
        usage = obj.get("usage")
        if usage:
            return normalize_usage(usage), None
        if obj.get("error"):
            return None, classify_error_response(body, status_code)
    except json.JSONDecodeError:
        pass

    # 流式 SSE
    result = extract_usage_from_sse(body)
    if result is not None:
        return result, None

    return None, classify_error_response(body, status_code)


def extract_usage_from_sse(text):
    """
    从流式 SSE 中提取 usage，需要合并多个事件：
    - message_start 包含 input_tokens, cache 等
    - message_delta 包含最终 output_tokens
    - 也可能来自 amazon-bedrock-invocationMetrics
    """
    merged = {}
    for line in text.split("\n"):
        line = line.strip()
        if not line.startswith("data:"):
            continue
        payload = line[len("data:"):].strip()
        if payload == "[DONE]":
            continue
        try:
            data = json.loads(payload)
        except json.JSONDecodeError:
            continue

        usage = None
        if "usage" in data and data["usage"]:
            usage = data["usage"]
        msg = data.get("message")
        if isinstance(msg, dict) and msg.get("usage"):
            usage = msg["usage"]

        if usage:
            _merge_usage(merged, usage)

        # Bedrock invocationMetrics 作为补充来源
        metrics = data.get("amazon-bedrock-invocationMetrics")
        if isinstance(metrics, dict):
            bedrock_usage = {}
            if "inputTokenCount" in metrics:
                bedrock_usage["input_tokens"] = metrics["inputTokenCount"]
            if "outputTokenCount" in metrics:
                bedrock_usage["output_tokens"] = metrics["outputTokenCount"]
            if "cacheReadInputTokenCount" in metrics:
                bedrock_usage["cache_read_input_tokens"] = metrics["cacheReadInputTokenCount"]
            if "cacheWriteInputTokenCount" in metrics:
                bedrock_usage["cache_creation_input_tokens"] = metrics["cacheWriteInputTokenCount"]
            if bedrock_usage:
                _merge_usage(merged, bedrock_usage)

    if merged:
        return normalize_usage(merged)
    return None


def _merge_usage(target, source):
    """将 source usage 合并到 target，对数值字段取较大值，对嵌套对象递归合并"""
    for k, v in source.items():
        if isinstance(v, dict):
            if k not in target or not isinstance(target[k], dict):
                target[k] = {}
            _merge_usage(target[k], v)
        elif isinstance(v, (int, float)):
            target[k] = max(target.get(k, 0), v)
        else:
            target[k] = v


def normalize_usage(usage):
    """提取标准化的 token 用量，包括 Claude 缓存 5m/1h 细分和工具调用"""
    cache_creation_total = int(usage.get("cache_creation_input_tokens") or 0)

    # Claude 缓存 5m/1h 细分：usage.cache_creation.ephemeral_5m/1h_input_tokens
    cache_creation_obj = usage.get("cache_creation")
    cache_5m = 0
    cache_1h = 0
    if isinstance(cache_creation_obj, dict):
        cache_5m = int(cache_creation_obj.get("ephemeral_5m_input_tokens") or 0)
        cache_1h = int(cache_creation_obj.get("ephemeral_1h_input_tokens") or 0)
        if cache_creation_total == 0 and (cache_5m or cache_1h):
            cache_creation_total = cache_5m + cache_1h

    # Web search 调用次数
    web_search_requests = 0
    server_tool_use = usage.get("server_tool_use")
    if isinstance(server_tool_use, dict):
        web_search_requests = int(server_tool_use.get("web_search_requests") or 0)

    return {
        "input_tokens": int(usage.get("input_tokens") or usage.get("prompt_tokens") or 0),
        "output_tokens": int(usage.get("output_tokens") or usage.get("completion_tokens") or 0),
        "cache_read_tokens": int(usage.get("cache_read_input_tokens") or 0),
        "cache_creation_tokens": cache_creation_total,
        "cache_creation_5m_tokens": cache_5m,
        "cache_creation_1h_tokens": cache_1h,
        "web_search_requests": web_search_requests,
    }


# ---------------------------------------------------------------------------
# 分段计费查找
# ---------------------------------------------------------------------------

def find_price_tier(model_pricing, input_tokens):
    """按 input_tokens 匹配分段价格区间，返回匹配的 tier dict 或 None"""
    tiers = model_pricing.get("tiered_pricing")
    if not tiers:
        return None

    input_tokens_k = input_tokens // 1000
    matched = None
    for tier in tiers:
        min_k = tier.get("min_tokens_k", 0)
        max_k = tier.get("max_tokens_k", -1)
        if input_tokens_k >= min_k:
            if max_k == -1 or input_tokens_k < max_k:
                matched = tier
                break

    # 未匹配则用最后一个区间（最高价格段）
    if matched is None and tiers:
        for t in reversed(tiers):
            if t.get("max_tokens_k", -1) == -1:
                return t
        return tiers[-1]

    return matched


def is_claude_model(model_name):
    return bool(CLAUDE_MODEL_RE.search(model_name))


# ---------------------------------------------------------------------------
# 费用计算
# ---------------------------------------------------------------------------

def calc_cost(usage, model_pricing, model_name, web_search_cfg):
    """
    根据 usage 和模型价格计算费用 (USD)。

    四条路径:
      0. 有 per_call_price -> 按次计费
      1. 有 tiered_pricing -> 按分段价格
      2. Claude 模型无 tiered_pricing -> 自动 200K 倍率
      3. 普通模型 -> 基础价格
    """
    # 路径 0：按次计费（视频/图片生成等）
    per_call = model_pricing.get("per_call_price")
    if per_call is not None:
        return per_call

    input_tokens = usage["input_tokens"]
    output_tokens = usage["output_tokens"]
    cache_read = usage["cache_read_tokens"]
    cache_creation_total = usage["cache_creation_tokens"]
    cache_5m = usage["cache_creation_5m_tokens"]
    cache_1h = usage["cache_creation_1h_tokens"]
    remaining_cache = max(cache_creation_total - cache_5m - cache_1h, 0)

    net_input = max(input_tokens - cache_read - cache_creation_total, 0)

    tier = find_price_tier(model_pricing, input_tokens)

    if tier is not None:
        # 路径 1：分段计费
        ip = tier.get("input_price", 0)
        op = tier.get("output_price", 0)
        chp = tier.get("cache_hit_price", 0)
        cwp = tier.get("cache_write_price", 0)
        cwp_1h = tier.get("cache_write_price_1h") or cwp

        cost = (
            net_input / 1_000_000 * ip
            + output_tokens / 1_000_000 * op
            + cache_read / 1_000_000 * chp
            + remaining_cache / 1_000_000 * cwp
            + cache_5m / 1_000_000 * cwp
            + cache_1h / 1_000_000 * cwp_1h
        )
    elif is_claude_model(model_name) and input_tokens >= CLAUDE_200K_THRESHOLD:
        # 路径 2：Claude 200K 自动倍率
        ip = model_pricing.get("input_price", 0) * CLAUDE_200K_INPUT_MULT
        op = model_pricing.get("output_price", 0) * CLAUDE_200K_OUTPUT_MULT
        chp = model_pricing.get("cache_hit_price", 0)
        cwp = model_pricing.get("cache_write_price", 0)
        cwp_1h = model_pricing.get("cache_write_price_1h") or cwp

        cost = (
            net_input / 1_000_000 * ip
            + output_tokens / 1_000_000 * op
            + cache_read / 1_000_000 * chp
            + remaining_cache / 1_000_000 * cwp
            + cache_5m / 1_000_000 * cwp
            + cache_1h / 1_000_000 * cwp_1h
        )
    else:
        # 路径 3：普通模型
        ip = model_pricing.get("input_price", 0)
        op = model_pricing.get("output_price", 0)
        chp = model_pricing.get("cache_hit_price", 0)
        cwp = model_pricing.get("cache_write_price", 0)
        cwp_1h = model_pricing.get("cache_write_price_1h") or cwp

        cost = (
            net_input / 1_000_000 * ip
            + output_tokens / 1_000_000 * op
            + cache_read / 1_000_000 * chp
            + remaining_cache / 1_000_000 * cwp
            + cache_5m / 1_000_000 * cwp
            + cache_1h / 1_000_000 * cwp_1h
        )

    # Web Search 附加费用
    ws_count = usage.get("web_search_requests", 0)
    if ws_count > 0 and web_search_cfg:
        ws_cost = calc_web_search_cost(model_name, ws_count, web_search_cfg)
        cost += ws_cost

    return cost


def calc_web_search_cost(model_name, call_count, web_search_cfg):
    """计算 Web Search 工具调用费用 (USD)"""
    if is_claude_model(model_name):
        price_per_k = web_search_cfg.get("claude", 10.0)
    elif (model_name.startswith("o3") or model_name.startswith("o4")
          or model_name.startswith("gpt-5")):
        price_per_k = web_search_cfg.get("openai_normal", 10.0)
    else:
        price_per_k = web_search_cfg.get("openai_high", 25.0)

    return call_count / 1000.0 * price_per_k


# ---------------------------------------------------------------------------
# 汇总维度
# ---------------------------------------------------------------------------

def get_group_key(record, usage, group_by):
    if group_by == "model":
        return record.get("model", "unknown")
    elif group_by == "channel":
        return f"{record.get('channel_id', 0)}-{record.get('channel_name', 'unknown')}"
    elif group_by == "user":
        return str(record.get("user_id", 0))
    elif group_by == "hour":
        try:
            dt = datetime.fromisoformat(record["created_at"].replace("Z", "+00:00"))
            return dt.strftime("%Y-%m-%d %H:00")
        except (ValueError, KeyError):
            return "unknown"
    return "unknown"


def new_stat_bucket():
    return {
        "count": 0,
        "input_tokens": 0,
        "output_tokens": 0,
        "cache_read_tokens": 0,
        "cache_creation_tokens": 0,
        "web_search_calls": 0,
        "cost": 0.0,
        "web_search_cost": 0.0,
        "errors": 0,
        "no_pricing": 0,
    }


def merge_stats(dst, src):
    for key, bucket in src.items():
        out = dst[key]
        for field, value in bucket.items():
            out[field] += value


class ProgressBar:
    def __init__(self, total, label, enabled=True, width=28):
        self.total = max(total, 0)
        self.label = label
        self.enabled = enabled and total > 1
        self.width = width
        self.completed = 0
        self.failures = 0
        self.start_time = time.time()
        self.last_draw = 0.0
        if self.enabled:
            self._draw(force=True)

    def update(self, completed=0, failures=0, force=False):
        self.completed += completed
        self.failures += failures
        if self.completed > self.total:
            self.completed = self.total
        self._draw(force=force)

    def close(self):
        if not self.enabled:
            return
        self._draw(force=True)
        if self.completed < self.total:
            print(file=sys.stderr, flush=True)

    def _draw(self, force=False):
        if not self.enabled:
            return
        now = time.time()
        if not force and self.completed < self.total and (now - self.last_draw) < 0.2:
            return
        self.last_draw = now
        elapsed = max(now - self.start_time, 0.001)
        pct = self.completed / max(self.total, 1)
        filled = min(self.width, int(self.width * pct))
        bar = "#" * filled + "-" * (self.width - filled)
        speed = self.completed / elapsed
        line = (
            f"\r  [{self.label}] [{bar}] "
            f"{self.completed}/{self.total} {pct * 100:5.1f}% "
            f"{speed:6.1f} 文件/s"
        )
        if self.failures:
            line += f"  失败:{self.failures}"
        end = "\n" if self.completed >= self.total else ""
        print(line, end=end, file=sys.stderr, flush=True)


def build_detail_row(date_str, record, model, usage, ws_count, cost, group_key):
    return {
        "date": date_str,
        "request_id": record.get("request_id", ""),
        "model": model,
        "channel_id": record.get("channel_id", 0),
        "channel_name": record.get("channel_name", ""),
        "user_id": record.get("user_id", 0),
        "status_code": record.get("status_code", 0),
        "input_tokens": usage["input_tokens"],
        "output_tokens": usage["output_tokens"],
        "cache_read_tokens": usage["cache_read_tokens"],
        "cache_creation_tokens": usage["cache_creation_tokens"],
        "cache_5m_tokens": usage["cache_creation_5m_tokens"],
        "cache_1h_tokens": usage["cache_creation_1h_tokens"],
        "web_search_calls": ws_count,
        "cost_usd": round(cost, 8),
        "group_key": group_key,
    }


def process_records(records, date_str, pricing_cfg, group_by,
                    filter_user_ids=None, filter_models=None, filter_channel_ids=None,
                    time_from=None, time_to=None, collect_details=True):
    models_pricing = pricing_cfg.get("models", {})
    web_search_cfg = pricing_cfg.get("web_search", {})
    user_id_set = set(filter_user_ids) if filter_user_ids else None
    model_set = set(filter_models) if filter_models else None
    channel_id_set = set(filter_channel_ids) if filter_channel_ids else None

    stats = defaultdict(new_stat_bucket)
    details = []
    total_records = len(records)
    filtered_out = 0
    parse_failures = 0
    error_categories = Counter()

    for rec in records:
        if time_from or time_to:
            ts = rec.get("created_at", "")
            if ts:
                if time_from and ts < time_from:
                    filtered_out += 1
                    continue
                if time_to and ts > time_to:
                    filtered_out += 1
                    continue

        if user_id_set and rec.get("user_id", 0) not in user_id_set:
            filtered_out += 1
            continue
        if channel_id_set and rec.get("channel_id", 0) not in channel_id_set:
            filtered_out += 1
            continue

        model = rec.get("model", "unknown")
        if model_set and model not in model_set:
            filtered_out += 1
            continue

        usage, fail_reason = extract_usage(rec)

        if usage is None:
            parse_failures += 1
            error_categories[fail_reason] += 1
            continue

        model_pricing = models_pricing.get(model)
        group_key = get_group_key(rec, usage, group_by)
        s = stats[group_key]
        s["count"] += 1
        s["input_tokens"] += usage["input_tokens"]
        s["output_tokens"] += usage["output_tokens"]
        s["cache_read_tokens"] += usage["cache_read_tokens"]
        s["cache_creation_tokens"] += usage["cache_creation_tokens"]
        s["web_search_calls"] += usage.get("web_search_requests", 0)

        if rec.get("status_code", 200) >= 400:
            s["errors"] += 1

        if model_pricing:
            cost = calc_cost(usage, model_pricing, model, web_search_cfg)
            s["cost"] += cost

            ws_count = usage.get("web_search_requests", 0)
            ws_cost = 0.0
            if ws_count > 0 and web_search_cfg:
                ws_cost = calc_web_search_cost(model, ws_count, web_search_cfg)
            s["web_search_cost"] += ws_cost

            if collect_details:
                details.append(build_detail_row(
                    date_str, rec, model, usage, ws_count, cost, group_key
                ))
        else:
            s["no_pricing"] += 1

    return stats, details, total_records, parse_failures, filtered_out, error_categories


def _chunk_keys(keys, chunk_size):
    for i in range(0, len(keys), chunk_size):
        yield keys[i:i + chunk_size]


def _process_key_batch(region, endpoint, bucket, keys, cache_dir, date_str, pricing_cfg, group_by,
                       filter_user_ids=None, filter_models=None, filter_channel_ids=None,
                       time_from=None, time_to=None, collect_details=True):
    client = get_s3_client(region, endpoint)
    stats = defaultdict(new_stat_bucket)
    details = []
    total_records = 0
    filtered_out = 0
    parse_failures = 0
    error_categories = Counter()
    download_errors = 0

    for key in keys:
        try:
            records = download_and_parse(client, bucket, key, cache_dir=cache_dir)
        except Exception as exc:
            download_errors += 1
            continue

        batch_stats, batch_details, batch_total, batch_parse_failures, batch_filtered_out, batch_errors = process_records(
            records, date_str, pricing_cfg, group_by,
            filter_user_ids=filter_user_ids, filter_models=filter_models,
            filter_channel_ids=filter_channel_ids, time_from=time_from, time_to=time_to,
            collect_details=collect_details,
        )
        merge_stats(stats, batch_stats)
        if collect_details and batch_details:
            details.extend(batch_details)
        total_records += batch_total
        filtered_out += batch_filtered_out
        parse_failures += batch_parse_failures
        error_categories.update(batch_errors)

    return {
        "stats": {k: dict(v) for k, v in stats.items()},
        "details": details,
        "total_records": total_records,
        "filtered_out": filtered_out,
        "parse_failures": parse_failures,
        "error_categories": dict(error_categories),
        "download_errors": download_errors,
    }


# ---------------------------------------------------------------------------
# 处理单天数据
# ---------------------------------------------------------------------------

def _download_one(region, endpoint, bucket, key, cache_dir=None):
    """在独立线程中创建 S3 client 并下载解析单个文件（避免 boto3 client 线程安全问题）"""
    if cache_dir:
        cache_path = _get_cache_path(cache_dir, key)
        if os.path.exists(cache_path):
            with open(cache_path, "rb") as f:
                return _parse_gzip_data(f.read())
    client = get_s3_client(region, endpoint)
    return download_and_parse(client, bucket, key, cache_dir=cache_dir)


def process_date(s3, bucket, prefix, date_str, pricing_cfg, group_by, verbose,
                 filter_user_ids=None, filter_models=None, filter_channel_ids=None,
                 workers=10, region="", endpoint="", cache_dir=None,
                 time_from=None, time_to=None, processes=0, collect_details=True):
    """处理单天数据，返回汇总结果和明细"""
    dt = datetime.strptime(date_str, "%Y-%m-%d")
    s3_prefix = f"{prefix}/{dt.strftime('%Y/%m/%d')}/"

    keys = list_s3_objects(s3, bucket, s3_prefix)
    if verbose:
        print(f"  [{date_str}] 找到 {len(keys)} 个文件")

    stats = defaultdict(new_stat_bucket)
    details = []
    total_records = 0
    filtered_out = 0
    parse_failures = 0
    error_categories = Counter()
    download_errors = 0

    t0 = time.time()

    # 统计缓存命中
    cache_hits = 0
    if cache_dir:
        for key in keys:
            if os.path.exists(_get_cache_path(cache_dir, key)):
                cache_hits += 1
        if verbose and cache_hits:
            print(f"  [{date_str}] 缓存命中: {cache_hits}/{len(keys)} 文件")
    progress = ProgressBar(len(keys), date_str)

    def merge_result(result):
        nonlocal total_records, filtered_out, parse_failures, download_errors
        merge_stats(stats, result["stats"])
        if collect_details and result["details"]:
            details.extend(result["details"])
        total_records += result["total_records"]
        filtered_out += result["filtered_out"]
        parse_failures += result["parse_failures"]
        error_categories.update(result["error_categories"])
        download_errors += result.get("download_errors", 0)

    if len(keys) <= 1:
        for key in keys:
            try:
                records = download_and_parse(s3, bucket, key, cache_dir=cache_dir)
            except Exception as e:
                download_errors += 1
                if verbose:
                    print(f"    下载失败 {key}: {e}")
                progress.update(completed=1, failures=1)
                continue
            batch_result = process_records(
                records, date_str, pricing_cfg, group_by,
                filter_user_ids=filter_user_ids, filter_models=filter_models,
                filter_channel_ids=filter_channel_ids, time_from=time_from, time_to=time_to,
                collect_details=collect_details,
            )
            batch_stats, batch_details, batch_total, batch_parse_failures, batch_filtered_out, batch_errors = batch_result
            merge_stats(stats, batch_stats)
            if collect_details and batch_details:
                details.extend(batch_details)
            total_records += batch_total
            filtered_out += batch_filtered_out
            parse_failures += batch_parse_failures
            error_categories.update(batch_errors)
            progress.update(completed=1)
    elif processes and processes > 1:
        actual_processes = min(processes, len(keys))
        chunk_size = max(200, min(1000, len(keys) // (actual_processes * 4) or 1))
        if verbose:
            print(f"  [{date_str}] 多进程处理: {actual_processes} 进程, 分块 {chunk_size} 文件")
        with ProcessPoolExecutor(max_workers=actual_processes) as pool:
            future_to_count = {
                pool.submit(
                    _process_key_batch,
                    region, endpoint, bucket, key_batch, cache_dir, date_str, pricing_cfg, group_by,
                    filter_user_ids, filter_models, filter_channel_ids,
                    time_from, time_to, collect_details,
                ): len(key_batch)
                for key_batch in _chunk_keys(keys, chunk_size)
            }
            for future in as_completed(future_to_count):
                batch_count = future_to_count[future]
                try:
                    result = future.result()
                    merge_result(result)
                    progress.update(completed=batch_count, failures=result.get("download_errors", 0))
                except Exception as e:
                    download_errors += batch_count
                    if verbose:
                        print(f"    分块处理失败: {e}")
                    progress.update(completed=batch_count, failures=batch_count)
    elif workers <= 1:
        for key in keys:
            try:
                records = download_and_parse(s3, bucket, key, cache_dir=cache_dir)
            except Exception as e:
                download_errors += 1
                if verbose:
                    print(f"    下载失败 {key}: {e}")
                progress.update(completed=1, failures=1)
                continue
            batch_result = process_records(
                records, date_str, pricing_cfg, group_by,
                filter_user_ids=filter_user_ids, filter_models=filter_models,
                filter_channel_ids=filter_channel_ids, time_from=time_from, time_to=time_to,
                collect_details=collect_details,
            )
            batch_stats, batch_details, batch_total, batch_parse_failures, batch_filtered_out, batch_errors = batch_result
            merge_stats(stats, batch_stats)
            if collect_details and batch_details:
                details.extend(batch_details)
            total_records += batch_total
            filtered_out += batch_filtered_out
            parse_failures += batch_parse_failures
            error_categories.update(batch_errors)
            progress.update(completed=1)
    else:
        actual_workers = min(workers, len(keys))
        with ThreadPoolExecutor(max_workers=actual_workers) as pool:
            future_to_key = {
                pool.submit(_download_one, region, endpoint, bucket, key, cache_dir): key
                for key in keys
            }
            for future in as_completed(future_to_key):
                key = future_to_key[future]
                try:
                    records = future.result()
                except Exception as e:
                    download_errors += 1
                    if verbose:
                        print(f"    下载失败 {key}: {e}")
                    progress.update(completed=1, failures=1)
                    continue
                batch_result = process_records(
                    records, date_str, pricing_cfg, group_by,
                    filter_user_ids=filter_user_ids, filter_models=filter_models,
                    filter_channel_ids=filter_channel_ids, time_from=time_from, time_to=time_to,
                    collect_details=collect_details,
                )
                batch_stats, batch_details, batch_total, batch_parse_failures, batch_filtered_out, batch_errors = batch_result
                merge_stats(stats, batch_stats)
                if collect_details and batch_details:
                    details.extend(batch_details)
                total_records += batch_total
                filtered_out += batch_filtered_out
                parse_failures += batch_parse_failures
                error_categories.update(batch_errors)
                progress.update(completed=1)

    elapsed = time.time() - t0
    progress.close()
    if verbose and len(keys) > 1:
        print(f"  [{date_str}] 下载完成: {len(keys)} 文件, {elapsed:.1f}s, "
              f"{len(keys)/max(elapsed,0.001):.0f} 文件/s"
              + (f", {download_errors} 失败" if download_errors else ""))

    return stats, details, total_records, parse_failures, filtered_out, error_categories


# ---------------------------------------------------------------------------
# 报告输出
# ---------------------------------------------------------------------------

def print_report(date_label, stats, total_records, parse_failures, group_by,
                 filtered_out=0, filters=None, error_categories=None):
    print(f"\n{'='*80}")
    print(f"  费用报告 - {date_label}")
    print(f"{'='*80}")
    summary = f"  日志总数: {total_records}    解析失败: {parse_failures}"
    if filtered_out:
        summary += f"    过滤掉: {filtered_out}"
    print(summary)
    if filters:
        print(f"  过滤条件: {filters}")

    if error_categories:
        print(f"\n  ── 失败原因归类 ({parse_failures} 条) ──")
        for reason, cnt in error_categories.most_common():
            print(f"    {cnt:>5}  {reason}")

    print()

    headers = [
        group_by.capitalize(), "请求数", "输入tokens", "输出tokens",
        "缓存读取", "缓存写入", "WebSearch", "错误", "无价格", "费用(USD)"
    ]
    rows = []
    total_cost = 0.0
    total_count = 0
    total_input = 0
    total_output = 0
    total_ws = 0

    for key in sorted(stats.keys()):
        s = stats[key]
        total_cost += s["cost"]
        total_count += s["count"]
        total_input += s["input_tokens"]
        total_output += s["output_tokens"]
        total_ws += s["web_search_calls"]
        rows.append([
            key,
            f"{s['count']:,}",
            f"{s['input_tokens']:,}",
            f"{s['output_tokens']:,}",
            f"{s['cache_read_tokens']:,}",
            f"{s['cache_creation_tokens']:,}",
            s["web_search_calls"] or "",
            s["errors"] or "",
            s["no_pricing"] or "",
            f"${s['cost']:.6f}",
        ])

    rows.append([
        "TOTAL",
        f"{total_count:,}",
        f"{total_input:,}",
        f"{total_output:,}",
        "", "",
        total_ws or "",
        "", "",
        f"${total_cost:.6f}",
    ])

    print(tabulate(rows, headers=headers, tablefmt="simple"))
    print(f"\n  总费用: ${total_cost:.6f} (约 ¥{total_cost * 7.3:.4f})")
    print(f"{'='*80}\n")


def export_csv(filepath, all_details, all_stats, group_by):
    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)

        w.writerow(["=== 汇总 ==="])
        w.writerow([group_by, "count", "input_tokens", "output_tokens",
                     "cache_read", "cache_write", "web_search_calls",
                     "errors", "no_pricing", "cost_usd", "web_search_cost_usd"])
        for key in sorted(all_stats.keys()):
            s = all_stats[key]
            w.writerow([key, s["count"], s["input_tokens"], s["output_tokens"],
                        s["cache_read_tokens"], s["cache_creation_tokens"],
                        s["web_search_calls"],
                        s["errors"], s["no_pricing"],
                        round(s["cost"], 8), round(s["web_search_cost"], 8)])

        w.writerow([])
        w.writerow(["=== 明细 ==="])
        if all_details:
            w.writerow(all_details[0].keys())
            for d in all_details:
                w.writerow(d.values())

    print(f"  CSV 已导出: {filepath}\n")


# ---------------------------------------------------------------------------
# Excel 账单导出
# ---------------------------------------------------------------------------

def export_bill(filepath, all_details, all_stats, date_label, args):
    """导出发票式 Excel 账单"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers

    wb = Workbook()
    ws = wb.active
    ws.title = "账单"

    currency = args.bill_currency
    rate = args.exchange_rate if currency == "CNY" else 1.0
    symbol = "¥" if currency == "CNY" else "$"
    currency_label = "人民币 (CNY)" if currency == "CNY" else "美元 (USD)"

    # 样式定义
    title_font = Font(name="微软雅黑", size=16, bold=True)
    header_font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    normal_font = Font(name="微软雅黑", size=10)
    money_font = Font(name="Consolas", size=10)
    total_font = Font(name="微软雅黑", size=11, bold=True)
    total_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="B0B0B0"),
        right=Side(style="thin", color="B0B0B0"),
        top=Side(style="thin", color="B0B0B0"),
        bottom=Side(style="thin", color="B0B0B0"),
    )
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    row = 1

    # ---- 标题 ----
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    cell = ws.cell(row=row, column=1, value=args.bill_title)
    cell.font = title_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    row += 2

    # ---- 账单信息 ----
    info_items = [
        ("账单周期", date_label),
        ("币种", currency_label),
        ("生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]
    if currency == "CNY":
        info_items.append(("汇率", f"1 USD = {rate} CNY"))
    if args.user_id:
        info_items.append(("用户 ID", ", ".join(str(u) for u in args.user_id)))
    if args.model:
        info_items.append(("模型", ", ".join(args.model)))
    if args.channel_id:
        info_items.append(("渠道 ID", ", ".join(str(c) for c in args.channel_id)))

    for label, value in info_items:
        ws.cell(row=row, column=1, value=label).font = Font(name="微软雅黑", size=10, bold=True)
        ws.cell(row=row, column=2, value=str(value)).font = normal_font
        row += 1
    row += 1

    # ---- 按模型汇总表 ----
    # 先按模型重新汇总 details
    model_summary = defaultdict(lambda: {
        "count": 0, "input_tokens": 0, "output_tokens": 0,
        "cache_read": 0, "cache_write": 0, "web_search": 0, "cost_usd": 0.0,
    })
    for d in all_details:
        m = d["model"]
        s = model_summary[m]
        s["count"] += 1
        s["input_tokens"] += d["input_tokens"]
        s["output_tokens"] += d["output_tokens"]
        s["cache_read"] += d["cache_read_tokens"]
        s["cache_write"] += d["cache_creation_tokens"]
        s["web_search"] += d.get("web_search_calls", 0)
        s["cost_usd"] += d["cost_usd"]

    headers = ["模型", "请求数", "输入 Tokens", "输出 Tokens",
               "缓存读取", "缓存写入", "WebSearch", f"费用 ({symbol})"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border
    row += 1

    grand_cost = 0.0
    for model_name in sorted(model_summary.keys()):
        ms = model_summary[model_name]
        cost_display = ms["cost_usd"] * rate
        grand_cost += cost_display
        values = [
            model_name, ms["count"], ms["input_tokens"], ms["output_tokens"],
            ms["cache_read"], ms["cache_write"], ms["web_search"],
            cost_display,
        ]
        for col_idx, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=v)
            cell.font = money_font if col_idx == 8 else normal_font
            cell.border = thin_border
            if col_idx == 1:
                cell.alignment = left
            elif col_idx == 8:
                cell.number_format = f'{symbol}#,##0.000000'
                cell.alignment = right
            else:
                cell.number_format = '#,##0'
                cell.alignment = right
        row += 1

    # 合计行
    for col_idx in range(1, 9):
        cell = ws.cell(row=row, column=col_idx)
        cell.fill = total_fill
        cell.font = total_font
        cell.border = thin_border
    ws.cell(row=row, column=1, value="合计").alignment = center
    ws.cell(row=row, column=2, value=sum(ms["count"] for ms in model_summary.values())).alignment = right
    ws.cell(row=row, column=2).number_format = '#,##0'
    cost_cell = ws.cell(row=row, column=8, value=grand_cost)
    cost_cell.number_format = f'{symbol}#,##0.000000'
    cost_cell.alignment = right
    cost_cell.font = Font(name="Consolas", size=11, bold=True)
    row += 2

    # ---- 请求明细表 ----
    if all_details:
        ws.cell(row=row, column=1, value="请求明细").font = Font(name="微软雅黑", size=12, bold=True)
        row += 1

        detail_headers = ["时间", "Request ID", "模型", "渠道", "用户ID",
                          "状态码", "输入", "输出", "缓存读取", "缓存写入",
                          f"费用 ({symbol})"]
        for col_idx, h in enumerate(detail_headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = thin_border
        row += 1

        for d in all_details:
            cost_d = d["cost_usd"] * rate
            values = [
                d["date"], d["request_id"], d["model"],
                f"{d['channel_id']}-{d['channel_name']}",
                d["user_id"], d["status_code"],
                d["input_tokens"], d["output_tokens"],
                d["cache_read_tokens"], d["cache_creation_tokens"],
                cost_d,
            ]
            for col_idx, v in enumerate(values, 1):
                cell = ws.cell(row=row, column=col_idx, value=v)
                cell.font = money_font if col_idx == 11 else normal_font
                cell.border = thin_border
                if col_idx == 11:
                    cell.number_format = f'{symbol}#,##0.000000'
                    cell.alignment = right
                elif col_idx >= 7:
                    cell.number_format = '#,##0'
                    cell.alignment = right
                else:
                    cell.alignment = left
            row += 1

    # ---- 列宽调整 ----
    col_widths = [22, 12, 14, 14, 12, 12, 12, 18, 12, 12, 18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else ""].width = w
    # 手动设置前几列
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 28
    ws.column_dimensions["H"].width = 18
    ws.column_dimensions["K"].width = 18

    wb.save(filepath)
    print(f"  Excel 账单已导出: {filepath}")
    print(f"  合计: {symbol}{grand_cost:,.6f}\n")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    args = parse_args()

    if not args.bucket:
        print("错误: 请设置 S3_BUCKET 或 --bucket", file=sys.stderr)
        sys.exit(1)

    script_dir = os.path.dirname(os.path.abspath(__file__))
    pricing_path = args.pricing if os.path.isabs(args.pricing) else os.path.join(script_dir, args.pricing)
    pricing_cfg = load_pricing(pricing_path)
    models_pricing = pricing_cfg.get("models", {})
    print(f"  已加载 {len(models_pricing)} 个模型价格配置")

    ws_cfg = pricing_cfg.get("web_search")
    if ws_cfg:
        print(f"  Web Search 计费: Claude ${ws_cfg.get('claude', 0)}/千次, "
              f"OpenAI ${ws_cfg.get('openai_high', 0)}/${ws_cfg.get('openai_normal', 0)}/千次")

    # 确定日期范围
    if args.date_range:
        start = datetime.strptime(args.date_range[0], "%Y-%m-%d")
        end = datetime.strptime(args.date_range[1], "%Y-%m-%d")
        dates = []
        cur = start
        while cur <= end:
            dates.append(cur.strftime("%Y-%m-%d"))
            cur += timedelta(days=1)
    elif args.date:
        dates = [args.date]
    else:
        dates = [(datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")]

    s3 = get_s3_client(args.region, args.endpoint)

    # 缓存目录
    cache_dir = None
    if not args.no_cache:
        cache_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), args.cache_dir)
        os.makedirs(cache_dir, exist_ok=True)
        if args.verbose:
            print(f"  缓存目录: {cache_dir}")

    # 时间过滤：短格式 HH:MM:SS 自动补全为当天的完整时间戳
    time_from = args.time_from
    time_to = args.time_to
    if time_from and len(time_from) <= 8:
        time_from = f"{dates[0]}T{time_from}"
    if time_to and len(time_to) <= 8:
        time_to = f"{dates[-1]}T{time_to}"

    # 构建过滤条件描述
    filter_parts = []
    if args.user_id:
        filter_parts.append(f"user_id={args.user_id}")
    if args.model:
        filter_parts.append(f"model={args.model}")
    if args.channel_id:
        filter_parts.append(f"channel_id={args.channel_id}")
    if time_from:
        filter_parts.append(f"from={time_from}")
    if time_to:
        filter_parts.append(f"to={time_to}")
    filters_desc = ", ".join(filter_parts) if filter_parts else None
    if filters_desc:
        print(f"  过滤条件: {filters_desc}")

    all_stats = defaultdict(new_stat_bucket)
    all_details = []
    grand_total_records = 0
    grand_parse_failures = 0
    grand_filtered_out = 0
    grand_error_categories = Counter()
    collect_details = bool(args.output or args.bill)

    if args.workers > 1:
        print(f"  并发下载: {args.workers} 线程")
    if args.processes > 1:
        print(f"  并发处理: {args.processes} 进程")
    if not collect_details:
        print("  明细导出: 已关闭（仅控制台汇总，减少内存占用）")

    for date_str in dates:
        stats, details, total_records, parse_failures, filtered_out, err_cats = process_date(
            s3, args.bucket, args.prefix, date_str, pricing_cfg, args.group_by, args.verbose,
            filter_user_ids=args.user_id, filter_models=args.model,
            filter_channel_ids=args.channel_id,
            workers=args.workers, region=args.region, endpoint=args.endpoint,
            cache_dir=cache_dir, time_from=time_from, time_to=time_to,
            processes=args.processes, collect_details=collect_details,
        )
        grand_total_records += total_records
        grand_parse_failures += parse_failures
        grand_filtered_out += filtered_out
        grand_error_categories += err_cats
        if collect_details and details:
            all_details.extend(details)

        merge_stats(all_stats, stats)

    date_label = dates[0] if len(dates) == 1 else f"{dates[0]} ~ {dates[-1]}"
    print_report(date_label, all_stats, grand_total_records, grand_parse_failures, args.group_by,
                 filtered_out=grand_filtered_out, filters=filters_desc,
                 error_categories=grand_error_categories)

    if args.output:
        export_csv(args.output, all_details, all_stats, args.group_by)

    if args.bill:
        export_bill(args.bill, all_details, all_stats, date_label, args)


if __name__ == "__main__":
    main()
