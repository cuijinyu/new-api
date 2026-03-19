import csv
from collections import defaultdict
from datetime import datetime

from tabulate import tabulate


def print_report(date_label, stats, total_records, parse_failures, group_by,
                 filtered_out=0, filters=None, error_categories=None):
    print(f"\n{'='*80}")
    print(f"  费用报告 - {date_label}")
    print(f"{'='*80}")
    summary = f"  日志总数: {total_records}    解析失败: {parse_failures}"
    if filtered_out:
        summary += f"    过滤掉: {filtered_out}"
    print(summary)
    if filters:
        print(f"  过滤条件: {filters}")

    if error_categories:
        print(f"\n  ── 失败原因归类 ({parse_failures} 条) ──")
        for reason, cnt in error_categories.most_common():
            print(f"    {cnt:>5}  {reason}")

    print()
    headers = [
        group_by.capitalize(), "请求数", "输入tokens", "输出tokens",
        "缓存读取", "缓存写入", "WebSearch", "错误", "无价格", "费用(USD)"
    ]
    rows = []
    total_cost = 0.0
    total_count = 0
    total_input = 0
    total_output = 0
    total_ws = 0

    for key in sorted(stats.keys()):
        s = stats[key]
        total_cost += s["cost"]
        total_count += s["count"]
        total_input += s["input_tokens"]
        total_output += s["output_tokens"]
        total_ws += s["web_search_calls"]
        rows.append([
            key,
            f"{s['count']:,}",
            f"{s['input_tokens']:,}",
            f"{s['output_tokens']:,}",
            f"{s['cache_read_tokens']:,}",
            f"{s['cache_creation_tokens']:,}",
            s["web_search_calls"] or "",
            s["errors"] or "",
            s["no_pricing"] or "",
            f"${s['cost']:.6f}",
        ])

    rows.append([
        "TOTAL",
        f"{total_count:,}",
        f"{total_input:,}",
        f"{total_output:,}",
        "", "",
        total_ws or "",
        "", "",
        f"${total_cost:.6f}",
    ])

    print(tabulate(rows, headers=headers, tablefmt="simple"))
    print(f"\n  总费用: ${total_cost:.6f} (约 ¥{total_cost * 7.3:.4f})")
    print(f"{'='*80}\n")


def export_csv(filepath, all_details, all_stats, group_by):
    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)

        w.writerow(["=== 汇总 ==="])
        w.writerow([group_by, "count", "input_tokens", "output_tokens",
                    "cache_read", "cache_write", "web_search_calls",
                    "errors", "no_pricing", "cost_usd", "web_search_cost_usd"])
        for key in sorted(all_stats.keys()):
            s = all_stats[key]
            w.writerow([key, s["count"], s["input_tokens"], s["output_tokens"],
                        s["cache_read_tokens"], s["cache_creation_tokens"],
                        s["web_search_calls"],
                        s["errors"], s["no_pricing"],
                        round(s["cost"], 8), round(s["web_search_cost"], 8)])

        w.writerow([])
        w.writerow(["=== 明细 ==="])
        if all_details:
            w.writerow(all_details[0].keys())
            for d in all_details:
                w.writerow(d.values())

    print(f"  CSV 已导出: {filepath}\n")


def export_bill(filepath, all_details, all_stats, date_label, args):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "账单"

    currency = args.bill_currency
    rate = args.exchange_rate if currency == "CNY" else 1.0
    symbol = "¥" if currency == "CNY" else "$"
    currency_label = "人民币 (CNY)" if currency == "CNY" else "美元 (USD)"

    title_font = Font(name="微软雅黑", size=16, bold=True)
    header_font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    normal_font = Font(name="微软雅黑", size=10)
    money_font = Font(name="Consolas", size=10)
    total_font = Font(name="微软雅黑", size=11, bold=True)
    total_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="B0B0B0"),
        right=Side(style="thin", color="B0B0B0"),
        top=Side(style="thin", color="B0B0B0"),
        bottom=Side(style="thin", color="B0B0B0"),
    )
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    cell = ws.cell(row=row, column=1, value=args.bill_title)
    cell.font = title_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    row += 2

    info_items = [
        ("账单周期", date_label),
        ("币种", currency_label),
        ("生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]
    if currency == "CNY":
        info_items.append(("汇率", f"1 USD = {rate} CNY"))
    if args.user_id:
        info_items.append(("用户 ID", ", ".join(str(u) for u in args.user_id)))
    if args.model:
        info_items.append(("模型", ", ".join(args.model)))
    if args.channel_id:
        info_items.append(("渠道 ID", ", ".join(str(c) for c in args.channel_id)))

    for label, value in info_items:
        ws.cell(row=row, column=1, value=label).font = Font(name="微软雅黑", size=10, bold=True)
        ws.cell(row=row, column=2, value=str(value)).font = normal_font
        row += 1
    row += 1

    model_summary = defaultdict(lambda: {
        "count": 0, "input_tokens": 0, "output_tokens": 0,
        "cache_read": 0, "cache_write": 0, "web_search": 0, "cost_usd": 0.0,
    })
    for d in all_details:
        m = d["model"]
        s = model_summary[m]
        s["count"] += 1
        s["input_tokens"] += d["input_tokens"]
        s["output_tokens"] += d["output_tokens"]
        s["cache_read"] += d["cache_read_tokens"]
        s["cache_write"] += d["cache_creation_tokens"]
        s["web_search"] += d.get("web_search_calls", 0)
        s["cost_usd"] += d["cost_usd"]

    headers = ["模型", "请求数", "输入 Tokens", "输出 Tokens",
               "缓存读取", "缓存写入", "WebSearch", f"费用 ({symbol})"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border
    row += 1

    grand_cost = 0.0
    for model_name in sorted(model_summary.keys()):
        ms = model_summary[model_name]
        cost_display = ms["cost_usd"] * rate
        grand_cost += cost_display
        values = [
            model_name, ms["count"], ms["input_tokens"], ms["output_tokens"],
            ms["cache_read"], ms["cache_write"], ms["web_search"], cost_display,
        ]
        for col_idx, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=v)
            cell.font = money_font if col_idx == 8 else normal_font
            cell.border = thin_border
            if col_idx == 1:
                cell.alignment = left
            elif col_idx == 8:
                cell.number_format = f"{symbol}#,##0.000000"
                cell.alignment = right
            else:
                cell.number_format = "#,##0"
                cell.alignment = right
        row += 1

    for col_idx in range(1, 9):
        cell = ws.cell(row=row, column=col_idx)
        cell.fill = total_fill
        cell.font = total_font
        cell.border = thin_border
    ws.cell(row=row, column=1, value="合计").alignment = center
    ws.cell(row=row, column=2, value=sum(ms["count"] for ms in model_summary.values())).alignment = right
    ws.cell(row=row, column=2).number_format = "#,##0"
    cost_cell = ws.cell(row=row, column=8, value=grand_cost)
    cost_cell.number_format = f"{symbol}#,##0.000000"
    cost_cell.alignment = right
    cost_cell.font = Font(name="Consolas", size=11, bold=True)
    row += 2

    if all_details:
        ws.cell(row=row, column=1, value="请求明细").font = Font(name="微软雅黑", size=12, bold=True)
        row += 1
        detail_headers = ["时间", "Request ID", "模型", "渠道", "用户ID",
                          "状态码", "输入", "输出", "缓存读取", "缓存写入", f"费用 ({symbol})"]
        for col_idx, h in enumerate(detail_headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = thin_border
        row += 1

        for d in all_details:
            cost_d = d["cost_usd"] * rate
            values = [
                d["date"], d["request_id"], d["model"],
                f"{d['channel_id']}-{d['channel_name']}",
                d["user_id"], d["status_code"],
                d["input_tokens"], d["output_tokens"],
                d["cache_read_tokens"], d["cache_creation_tokens"], cost_d,
            ]
            for col_idx, v in enumerate(values, 1):
                cell = ws.cell(row=row, column=col_idx, value=v)
                cell.font = money_font if col_idx == 11 else normal_font
                cell.border = thin_border
                if col_idx == 11:
                    cell.number_format = f"{symbol}#,##0.000000"
                    cell.alignment = right
                elif col_idx >= 7:
                    cell.number_format = "#,##0"
                    cell.alignment = right
                else:
                    cell.alignment = left
            row += 1

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 28
    ws.column_dimensions["H"].width = 18
    ws.column_dimensions["K"].width = 18

    wb.save(filepath)
    print(f"  Excel 账单已导出: {filepath}")
    print(f"  合计: {symbol}{grand_cost:,.6f}\n")
