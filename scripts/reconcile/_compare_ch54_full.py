"""
渠道54全量 vs 知书万卷账单 完整对账
不限用户，用降档逻辑重算 expected_usd
"""
import sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
os.chdir(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
import sqlite3
import json
import pandas as pd
import numpy as np
from collections import defaultdict
from datetime import datetime, timezone, timedelta

from gen_bill import (
    parse_other_batch, assign_prices, compute_costs,
    FLAT_TIER_MODELS, QUOTA_PER_USD, DB_PATH
)

BJ = timezone(timedelta(hours=8))
BILL_PATH = "reconcile/aws5101消费记录（2026-03-19 00-00-00 - 2026-03-25 00-00-00）.xlsx"
FLAT_TIER_SINCE_TS = int(datetime(2026, 3, 13, tzinfo=BJ).timestamp())

def ts_to_date(ts):
    return datetime.fromtimestamp(ts, tz=BJ).strftime("%m-%d")

# ── 1. 解析知书万卷账单 ──
print("Loading their bill...")
wb = openpyxl.load_workbook(BILL_PATH, read_only=True, data_only=True)
ws = wb[wb.sheetnames[0]]

their_rows = []
ts_min, ts_max = float("inf"), 0
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0:
        continue
    rid, user_id, created_at, typ = row[0], row[1], row[2], row[3]
    model_name, quota, prompt_tokens, completion_tokens = row[7], row[8], row[9], row[10]
    if typ != 2:
        continue
    their_rows.append({
        "id": rid, "created_at": created_at,
        "model": model_name,
        "quota": quota or 0,
        "prompt_tokens": prompt_tokens or 0,
        "completion_tokens": completion_tokens or 0,
        "date": ts_to_date(created_at),
    })
    ts_min = min(ts_min, created_at)
    ts_max = max(ts_max, created_at)
wb.close()

their_df = pd.DataFrame(their_rows)
print(f"Their total: {len(their_df):,} rows, ts: {datetime.fromtimestamp(ts_min, tz=BJ)} ~ {datetime.fromtimestamp(ts_max, tz=BJ)}")

# ── 2. 从数据库加载渠道54全量 ──
print("Loading our ch=54 data...")
conn = sqlite3.connect(DB_PATH)
our_raw = pd.read_sql_query(
    """SELECT id, user_id, username, model_name AS model, quota, prompt_tokens, completion_tokens,
              other, token_name, created_at,
              datetime(created_at,'unixepoch','+8 hours') AS cst
       FROM logs
       WHERE channel_id=54 AND type=2
         AND created_at>=? AND created_at<=?
       ORDER BY created_at""",
    conn, params=(ts_min, ts_max)
)
conn.close()
print(f"Our total: {len(our_raw):,} rows")

our_raw["date"] = our_raw["created_at"].apply(ts_to_date)

# ── 3. 降档重算 ──
print("Computing with flat-tier pricing...")
other_df = parse_other_batch(our_raw["other"])
df = pd.concat([our_raw.drop(columns=["other"]), other_df], axis=1)
df = assign_prices(df, flat_tier=True, flat_tier_since_ts=FLAT_TIER_SINCE_TS)
df = df.dropna(subset=["ip"])
df = compute_costs(df, flat_tier=True, flat_tier_since_ts=FLAT_TIER_SINCE_TS)

# ── 4. 汇总 ──
# 知书万卷按模型
t_model = their_df.groupby("model").agg(
    t_count=("id", "count"), t_quota=("quota", "sum"),
    t_prompt=("prompt_tokens", "sum"), t_comp=("completion_tokens", "sum"),
).reset_index()
t_model["t_usd"] = t_model["t_quota"] / QUOTA_PER_USD

# 我方按模型
o_model = df.groupby("model").agg(
    o_count=("id", "count"),
    o_billed=("billed_usd", "sum"),
    o_expected=("expected_usd", "sum"),
    o_prompt=("prompt_tokens", "sum"),
    o_comp=("completion_tokens", "sum"),
).reset_index()

cmp = t_model.merge(o_model, on="model", how="outer").fillna(0)
cmp["diff_exp"] = cmp["o_expected"] - cmp["t_usd"]
cmp["diff_pct"] = cmp["diff_exp"] / cmp["t_usd"].replace(0, np.nan) * 100
cmp = cmp.sort_values("t_usd", ascending=False)

print(f"\n{'='*150}")
print(f"渠道54全量 vs 知书万卷账单 (3/19~3/24, 所有用户)")
print(f"{'='*150}")
print(f"{'model':<35} {'t_cnt':>10} {'o_cnt':>10} {'d_cnt':>8} {'t_prompt':>14} {'o_prompt':>14} {'t_$(quota)':>14} {'o_billed$':>14} {'o_expected$':>14} {'diff(exp)':>12} {'%':>7}")
print("-" * 150)
for _, r in cmp.iterrows():
    d_cnt = int(r["o_count"] - r["t_count"])
    pct = r["diff_pct"] if pd.notna(r["diff_pct"]) else 0
    print(f"  {r['model']:<33} {int(r['t_count']):>10,} {int(r['o_count']):>10,} {d_cnt:>+8,} "
          f"{int(r['t_prompt']):>14,} {int(r['o_prompt']):>14,} "
          f"{r['t_usd']:>14,.2f} {r['o_billed']:>14,.2f} {r['o_expected']:>14,.2f} "
          f"{r['diff_exp']:>+12,.2f} {pct:>+6.1f}%")
# totals
print("-" * 150)
st = cmp.sum(numeric_only=True)
total_diff = st["o_expected"] - st["t_usd"]
total_pct = total_diff / st["t_usd"] * 100 if st["t_usd"] else 0
print(f"  {'TOTAL':<33} {int(st['t_count']):>10,} {int(st['o_count']):>10,} {int(st['o_count']-st['t_count']):>+8,} "
      f"{int(st['t_prompt']):>14,} {int(st['o_prompt']):>14,} "
      f"{st['t_usd']:>14,.2f} {st['o_billed']:>14,.2f} {st['o_expected']:>14,.2f} "
      f"{total_diff:>+12,.2f} {total_pct:>+6.1f}%")

# ── 5. 按日期对比 ──
t_daily = their_df.groupby("date").agg(t_count=("id","count"), t_quota=("quota","sum")).reset_index()
t_daily["t_usd"] = t_daily["t_quota"] / QUOTA_PER_USD

o_daily = df.groupby("date").agg(
    o_count=("id","count"), o_billed=("billed_usd","sum"), o_expected=("expected_usd","sum"),
).reset_index()

d_cmp = t_daily.merge(o_daily, on="date", how="outer").fillna(0)
d_cmp = d_cmp.sort_values("date")

print(f"\n{'='*120}")
print("按日期对比 (所有模型合计)")
print(f"{'='*120}")
print(f"{'date':<8} {'t_cnt':>10} {'o_cnt':>10} {'d_cnt':>8} {'t_$(quota)':>14} {'o_billed$':>14} {'o_expected$':>14} {'diff(exp)':>12} {'%':>7}")
print("-" * 100)
for _, r in d_cmp.iterrows():
    diff = r["o_expected"] - r["t_usd"]
    pct = diff / r["t_usd"] * 100 if r["t_usd"] else 0
    print(f"  {r['date']:<6} {int(r['t_count']):>10,} {int(r['o_count']):>10,} {int(r['o_count']-r['t_count']):>+8,} "
          f"{r['t_usd']:>14,.2f} {r['o_billed']:>14,.2f} {r['o_expected']:>14,.2f} "
          f"{diff:>+12,.2f} {pct:>+6.1f}%")

# ── 6. 按日期+模型 重点模型 ──
for focus in ["claude-opus-4-6", "claude-sonnet-4-6", "claude-opus-4-5-20251101"]:
    t_dm = their_df[their_df["model"]==focus].groupby("date").agg(
        t_count=("id","count"), t_quota=("quota","sum"),
    ).reset_index()
    t_dm["t_usd"] = t_dm["t_quota"] / QUOTA_PER_USD

    o_dm = df[df["model"]==focus].groupby("date").agg(
        o_count=("id","count"), o_billed=("billed_usd","sum"), o_expected=("expected_usd","sum"),
    ).reset_index()

    dm = t_dm.merge(o_dm, on="date", how="outer").fillna(0).sort_values("date")

    print(f"\n{'='*120}")
    print(f"按日期对比: {focus}")
    print(f"{'='*120}")
    print(f"{'date':<8} {'t_cnt':>10} {'o_cnt':>10} {'d_cnt':>8} {'t_$(quota)':>14} {'o_billed$':>14} {'o_expected$':>14} {'diff(exp)':>12} {'%':>7}")
    print("-" * 100)
    for _, r in dm.iterrows():
        diff = r["o_expected"] - r["t_usd"]
        pct = diff / r["t_usd"] * 100 if r["t_usd"] else 0
        print(f"  {r['date']:<6} {int(r['t_count']):>10,} {int(r['o_count']):>10,} {int(r['o_count']-r['t_count']):>+8,} "
              f"{r['t_usd']:>14,.2f} {r['o_billed']:>14,.2f} {r['o_expected']:>14,.2f} "
              f"{diff:>+12,.2f} {pct:>+6.1f}%")

# ── 7. 我方按用户分 ──
print(f"\n{'='*120}")
print("我方 ch=54 按用户分 (降档后)")
print(f"{'='*120}")
u_grp = df.groupby(["user_id", "username"]).agg(
    count=("id","count"), billed=("billed_usd","sum"), expected=("expected_usd","sum"),
).reset_index().sort_values("expected", ascending=False)
print(f"{'user_id':>8} {'username':<15} {'count':>10} {'billed$':>14} {'expected$':>14}")
print("-" * 70)
for _, r in u_grp.iterrows():
    print(f"  {int(r['user_id']):>6} {r['username']:<15} {int(r['count']):>10,} {r['billed']:>14,.2f} {r['expected']:>14,.2f}")
print("-" * 70)
print(f"  {'':>6} {'TOTAL':<15} {int(u_grp['count'].sum()):>10,} {u_grp['billed'].sum():>14,.2f} {u_grp['expected'].sum():>14,.2f}")
print(f"\n  知书万卷账单总计: {len(their_df):,} calls, ${their_df['quota'].sum()/QUOTA_PER_USD:,.2f}")

# ── 8. 次数差异分析 ──
print(f"\n{'='*120}")
print("次数差异分析")
print(f"{'='*120}")
our_total = len(df)
their_total = len(their_df)
print(f"知书万卷: {their_total:,} calls")
print(f"我方 ch=54: {our_total:,} calls")
print(f"差异: {our_total - their_total:+,} calls")

# quota=0 的记录
q0 = df[df["quota"] == 0]
print(f"\n我方 quota=0 记录: {len(q0):,} (这些在知书万卷可能不计费)")
q0_model = q0.groupby("model").size().sort_values(ascending=False)
for m, c in q0_model.items():
    print(f"  {m}: {c:,}")

# 被 dropna 过滤的记录（不在 PRICING 中的模型）
dropped = len(our_raw) - len(df)
if dropped > 0:
    print(f"\n被 PRICING 过滤的记录: {dropped:,}")
    missing_models = our_raw[~our_raw["model"].isin(df["model"].unique())]["model"].value_counts()
    for m, c in missing_models.items():
        print(f"  {m}: {c:,}")
