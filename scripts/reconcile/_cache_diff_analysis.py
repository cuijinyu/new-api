"""
深入对比知书万卷 vs 我方 cache 计费差异
重点分析 sonnet-4-6 和 opus-4-6
"""
import sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
os.chdir(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
import sqlite3
import json
import pandas as pd
import numpy as np
from collections import defaultdict
from datetime import datetime, timezone, timedelta

from gen_bill import DB_PATH, QUOTA_PER_USD

BJ = timezone(timedelta(hours=8))
BILL_PATH = "reconcile/aws5101消费记录（2026-03-19 00-00-00 - 2026-03-25 00-00-00）.xlsx"

def ts_to_date(ts):
    return datetime.fromtimestamp(ts, tz=BJ).strftime("%m-%d")

def safe_json(s):
    try:
        return json.loads(s) if s else {}
    except:
        return {}

# ── 1. 解析知书万卷账单 other 字段 ──
print("Loading their bill...")
wb = openpyxl.load_workbook(BILL_PATH, read_only=True, data_only=True)
ws = wb[wb.sheetnames[0]]

their_rows = []
ts_min, ts_max = float("inf"), 0
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0:
        continue
    rid, user_id, created_at, typ = row[0], row[1], row[2], row[3]
    model_name, quota, prompt_tokens, completion_tokens = row[7], row[8], row[9], row[10]
    other_str = row[-1] if len(row) >= 18 else ""
    if typ != 2:
        continue
    o = safe_json(other_str)
    their_rows.append({
        "id": rid, "created_at": created_at, "model": model_name,
        "quota": quota or 0,
        "prompt_tokens": prompt_tokens or 0,
        "completion_tokens": completion_tokens or 0,
        "t_cache_hit": o.get("cache_tokens", 0) or 0,
        "t_cache_write": o.get("cache_creation_tokens", 0) or 0,
        "t_model_ratio": o.get("model_ratio", 0),
        "t_comp_ratio": o.get("completion_ratio", 0),
        "t_cache_ratio": o.get("cache_ratio", 0),
        "t_cache_creation_ratio": o.get("cache_creation_ratio", 0),
        "date": ts_to_date(created_at),
    })
    ts_min = min(ts_min, created_at)
    ts_max = max(ts_max, created_at)
wb.close()
their_df = pd.DataFrame(their_rows)
print(f"Their: {len(their_df):,} rows")

# ── 2. 我方数据 ──
print("Loading our ch=54 data...")
conn = sqlite3.connect(DB_PATH)
our_raw = pd.read_sql_query(
    """SELECT id, model_name AS model, quota, prompt_tokens, completion_tokens,
              other, created_at,
              datetime(created_at,'unixepoch','+8 hours') AS cst
       FROM logs
       WHERE channel_id=54 AND type=2
         AND created_at>=? AND created_at<=?
       ORDER BY created_at""",
    conn, params=(ts_min, ts_max)
)
conn.close()
print(f"Ours: {len(our_raw):,} rows")

# 解析我方 other
def parse_our_other(s):
    o = safe_json(s)
    ch = o.get("cache_tokens", 0) or 0
    cw = o.get("cache_creation_tokens", 0) or 0
    cw_5m = o.get("tiered_cache_creation_tokens_5m") or o.get("cache_creation_tokens_5m", 0) or 0
    cw_1h = o.get("tiered_cache_creation_tokens_1h") or o.get("cache_creation_tokens_1h", 0) or 0
    cw_rem_raw = o.get("tiered_cache_creation_tokens_remaining")
    cw_rem = cw_rem_raw if cw_rem_raw is not None else max(cw - cw_5m - cw_1h, 0)
    return pd.Series({
        "o_cache_hit": ch,
        "o_cache_write_total": cw,
        "o_cw_5m": cw_5m,
        "o_cw_1h": cw_1h,
        "o_cw_rem": cw_rem,
        "o_has_tiered": 1 if "tiered_cache_store_price" in o else 0,
    })

our_cache = our_raw["other"].apply(parse_our_other)
our_df = pd.concat([our_raw.drop(columns=["other"]), our_cache], axis=1)
our_df["date"] = our_df["created_at"].apply(ts_to_date)

# ── 3. 按模型对比 cache 总量 ──
for focus in ["claude-sonnet-4-6", "claude-opus-4-6", "claude-opus-4-5-20251101"]:
    t_sub = their_df[their_df["model"] == focus]
    o_sub = our_df[our_df["model"] == focus]

    print(f"\n{'='*130}")
    print(f"Cache 对比: {focus}")
    print(f"{'='*130}")

    # 总量
    t_ch = t_sub["t_cache_hit"].sum()
    t_cw = t_sub["t_cache_write"].sum()
    o_ch = o_sub["o_cache_hit"].sum()
    o_cw = o_sub["o_cache_write_total"].sum()
    o_cw5m = o_sub["o_cw_5m"].sum()
    o_cw1h = o_sub["o_cw_1h"].sum()
    o_cwrem = o_sub["o_cw_rem"].sum()

    print(f"\n  总量对比:")
    print(f"  {'':30} {'知书万卷':>18} {'我方':>18} {'差异':>18} {'差%':>8}")
    print(f"  {'cache_hit tokens':<30} {t_ch:>18,} {o_ch:>18,} {o_ch-t_ch:>+18,} {(o_ch-t_ch)/max(t_ch,1)*100:>+7.1f}%")
    print(f"  {'cache_write tokens':<30} {t_cw:>18,} {o_cw:>18,} {o_cw-t_cw:>+18,} {(o_cw-t_cw)/max(t_cw,1)*100:>+7.1f}%")
    print(f"  {'  cw_5m (rem+5m)':<30} {'':>18} {o_cwrem+o_cw5m:>18,}")
    print(f"  {'  cw_1h':<30} {'':>18} {o_cw1h:>18,}")

    # 按日期
    t_daily = t_sub.groupby("date").agg(
        t_cnt=("id","count"), t_ch=("t_cache_hit","sum"), t_cw=("t_cache_write","sum"), t_quota=("quota","sum"),
    ).reset_index()
    o_daily = o_sub.groupby("date").agg(
        o_cnt=("id","count"), o_ch=("o_cache_hit","sum"), o_cw=("o_cache_write_total","sum"),
        o_cw5m=("o_cw_5m","sum"), o_cw1h=("o_cw_1h","sum"), o_cwrem=("o_cw_rem","sum"),
        o_quota=("quota","sum"),
    ).reset_index()
    daily = t_daily.merge(o_daily, on="date", how="outer").fillna(0).sort_values("date")

    print(f"\n  按日期:")
    print(f"  {'date':<8} {'t_ch':>14} {'o_ch':>14} {'ch_diff':>14} {'t_cw':>14} {'o_cw':>14} {'cw_diff':>14} {'t_$(q)':>12} {'o_$(q)':>12}")
    print(f"  {'-'*120}")
    for _, r in daily.iterrows():
        print(f"  {r['date']:<6} {int(r['t_ch']):>14,} {int(r['o_ch']):>14,} {int(r['o_ch']-r['t_ch']):>+14,} "
              f"{int(r['t_cw']):>14,} {int(r['o_cw']):>14,} {int(r['o_cw']-r['t_cw']):>+14,} "
              f"{r['t_quota']/QUOTA_PER_USD:>12,.2f} {r['o_quota']/QUOTA_PER_USD:>12,.2f}")

    # 知书万卷的 ratio 分布
    print(f"\n  知书万卷 ratio 分布:")
    for col in ["t_model_ratio", "t_comp_ratio", "t_cache_ratio", "t_cache_creation_ratio"]:
        vc = t_sub[col].value_counts().head(5)
        vals = ", ".join(f"{v}({c:,})" for v, c in vc.items())
        print(f"    {col}: {vals}")

    # 我方 tiered cache 比例
    tiered_pct = o_sub["o_has_tiered"].mean() * 100
    print(f"  我方 tiered cache 记录占比: {tiered_pct:.1f}%")

# ── 4. sonnet-4-6 逐条抽样对比 ──
print(f"\n\n{'='*130}")
print("sonnet-4-6 逐条抽样: cache 差异最大的记录")
print(f"{'='*130}")

# 取 sonnet-4-6 有 cache 的记录，按 quota 从大到小
t_sonnet = their_df[(their_df["model"] == "claude-sonnet-4-6") & (their_df["t_cache_hit"] > 0)].copy()
t_sonnet = t_sonnet.sort_values("quota", ascending=False).head(30)

print(f"\n  知书万卷 sonnet-4-6 TOP30 (有cache, 按quota排序):")
print(f"  {'id':>12} {'date':<6} {'prompt':>10} {'comp':>10} {'cache_hit':>12} {'cache_write':>12} {'quota':>12} {'$(q)':>10} {'mr':>4} {'cr':>4} {'chr':>5} {'cwr':>5}")
print(f"  {'-'*110}")
for _, r in t_sonnet.iterrows():
    print(f"  {int(r['id']):>12} {r['date']:<6} {int(r['prompt_tokens']):>10,} {int(r['completion_tokens']):>10,} "
          f"{int(r['t_cache_hit']):>12,} {int(r['t_cache_write']):>12,} "
          f"{int(r['quota']):>12,} {r['quota']/QUOTA_PER_USD:>10,.4f} "
          f"{r['t_model_ratio']:>4} {r['t_comp_ratio']:>4} {r['t_cache_ratio']:>5} {r['t_cache_creation_ratio']:>5}")

# ── 5. 验证知书万卷 quota 计算公式 ──
print(f"\n\n{'='*130}")
print("验证知书万卷 quota 计算公式 (sonnet-4-6, 有cache的记录)")
print(f"{'='*130}")

# 尝试多种公式
t_with_cache = their_df[(their_df["model"] == "claude-sonnet-4-6") & 
                         ((their_df["t_cache_hit"] > 0) | (their_df["t_cache_write"] > 0))].copy()
print(f"有 cache 的 sonnet-4-6 记录: {len(t_with_cache):,}")

# 公式: quota = (net_input*mr + comp*mr*cr + ch*mr*chr + cw*mr*cwr) * group_multiplier
# group_multiplier 可能是 2 (因为 quota 基数是 500000/USD)
for _, r in t_with_cache.head(20).iterrows():
    pt = int(r["prompt_tokens"])
    ct = int(r["completion_tokens"])
    ch = int(r["t_cache_hit"])
    cw = int(r["t_cache_write"])
    mr = r["t_model_ratio"]
    cr = r["t_comp_ratio"]
    chr_ = r["t_cache_ratio"]
    cwr = r["t_cache_creation_ratio"]
    actual = int(r["quota"])
    
    net = max(pt - ch - cw, 0)
    
    # 公式1: (net*mr + comp*mr*cr + ch*mr*chr + cw*mr*cwr) * 2
    calc1 = (net * mr + ct * mr * cr + ch * mr * chr_ + cw * mr * cwr) * 2
    # 公式2: 不乘2
    calc2 = net * mr + ct * mr * cr + ch * mr * chr_ + cw * mr * cwr
    # 公式3: 用 prompt (不减cache) 
    calc3 = (pt * mr + ct * mr * cr + ch * mr * chr_ + cw * mr * cwr) * 2
    # 公式4: net_input 不用 mr
    calc4 = (net + ct * cr + ch * chr_ + cw * cwr) * mr * 2
    
    match = ""
    for name, val in [("f1", calc1), ("f2", calc2), ("f3", calc3), ("f4", calc4)]:
        if abs(val - actual) < 2:
            match = name
            break
    
    print(f"  id={int(r['id']):>10} pt={pt:>8,} ct={ct:>8,} ch={ch:>10,} cw={cw:>10,} "
          f"mr={mr} cr={cr} chr={chr_} cwr={cwr} "
          f"actual={actual:>12,} f1={calc1:>12,.0f} f2={calc2:>12,.0f} "
          f"{'MATCH:'+match if match else 'NO MATCH  ratio='+f'{actual/max(calc2,1):.2f}'}")

# ── 6. 对比无 cache 的记录 ──
print(f"\n\n{'='*130}")
print("sonnet-4-6 无 cache 记录的 quota 验证")
print(f"{'='*130}")

t_no_cache = their_df[(their_df["model"] == "claude-sonnet-4-6") & 
                       (their_df["t_cache_hit"] == 0) & (their_df["t_cache_write"] == 0)].copy()
print(f"无 cache 的 sonnet-4-6 记录: {len(t_no_cache):,}")

# 这些记录 quota 应该 = prompt * mr + comp * mr * cr (某个倍数)
for _, r in t_no_cache.head(20).iterrows():
    pt = int(r["prompt_tokens"])
    ct = int(r["completion_tokens"])
    mr = r["t_model_ratio"]
    cr = r["t_comp_ratio"]
    actual = int(r["quota"])
    
    calc_base = pt * mr + ct * mr * cr
    calc_x2 = calc_base * 2
    
    if calc_base > 0:
        ratio = actual / calc_base
    else:
        ratio = 0
    
    match = "x2" if abs(calc_x2 - actual) < 2 else ("x1" if abs(calc_base - actual) < 2 else "")
    
    print(f"  id={int(r['id']):>10} pt={pt:>8,} ct={ct:>8,} mr={mr} cr={cr} "
          f"actual={actual:>12,} base={calc_base:>10,.0f} x2={calc_x2:>10,.0f} "
          f"{'MATCH:'+match if match else f'actual/base={ratio:.1f}'}")

# ── 7. 看看知书万卷 other 字段里还有什么我们没注意到的 key ──
print(f"\n\n{'='*130}")
print("知书万卷 other 字段 key 分布 (sonnet-4-6 抽样)")
print(f"{'='*130}")

wb = openpyxl.load_workbook(BILL_PATH, read_only=True, data_only=True)
ws = wb[wb.sheetnames[0]]
key_counts = defaultdict(int)
sample_others = []
count = 0
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0:
        continue
    if row[7] != "claude-sonnet-4-6" or row[3] != 2:
        continue
    other_str = row[-1] if len(row) >= 18 else ""
    o = safe_json(other_str)
    for k in o.keys():
        key_counts[k] += 1
    if count < 5:
        sample_others.append(o)
    count += 1
wb.close()

print(f"Total sonnet-4-6 records scanned: {count:,}")
print(f"\nAll keys in other:")
for k, c in sorted(key_counts.items(), key=lambda x: -x[1]):
    print(f"  {k}: {c:,} ({c/count*100:.1f}%)")

print(f"\nSample other (first 3):")
for i, o in enumerate(sample_others[:3]):
    print(f"  [{i}] {json.dumps(o, ensure_ascii=False)}")
