"""对比知书万卷账单 vs 我方数据库（按日期+模型维度）"""
import openpyxl
import sqlite3
from collections import defaultdict
from datetime import datetime, timezone, timedelta

BILL_PATH = "reconcile/aws5101消费记录（2026-03-19 00-00-00 - 2026-03-25 00-00-00）.xlsx"
DB_PATH = "logs_analysis.db"
QUOTA_PER_USD = 500_000
BJ = timezone(timedelta(hours=8))

def ts_to_date(ts):
    return datetime.fromtimestamp(ts, tz=BJ).strftime("%m-%d")

# ── 1. 解析知书万卷账单 ──
wb = openpyxl.load_workbook(BILL_PATH, read_only=True, data_only=True)
ws = wb[wb.sheetnames[0]]

their = defaultdict(lambda: defaultdict(lambda: {"count": 0, "quota": 0}))
their_total_model = defaultdict(lambda: {"count": 0, "quota": 0})
ts_min, ts_max = float("inf"), 0

for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0:
        continue
    rid, user_id, created_at, typ, content, username, token_name, model_name, quota = row[:9]
    if typ != 2:
        continue
    q = quota or 0
    date = ts_to_date(created_at)
    their[date][model_name]["count"] += 1
    their[date][model_name]["quota"] += q
    their_total_model[model_name]["count"] += 1
    their_total_model[model_name]["quota"] += q
    ts_min = min(ts_min, created_at)
    ts_max = max(ts_max, created_at)

wb.close()

# ── 2. 查我方数据 ──
conn = sqlite3.connect(DB_PATH)
rows = conn.execute("""
    SELECT created_at, model_name, quota
    FROM logs
    WHERE channel_id=54 AND type=2
      AND created_at>=? AND created_at<=?
""", (ts_min, ts_max)).fetchall()
conn.close()

ours = defaultdict(lambda: defaultdict(lambda: {"count": 0, "quota": 0}))
ours_total_model = defaultdict(lambda: {"count": 0, "quota": 0})
for created_at, model_name, quota in rows:
    q = quota or 0
    date = ts_to_date(created_at)
    ours[date][model_name]["count"] += 1
    ours[date][model_name]["quota"] += q
    ours_total_model[model_name]["count"] += 1
    ours_total_model[model_name]["quota"] += q

# ── 3. 汇总对比 ──
all_models = sorted(set(list(their_total_model.keys()) + list(ours_total_model.keys())))
all_dates = sorted(set(list(their.keys()) + list(ours.keys())))

print("=" * 120)
print("知书万卷账单 vs 我方数据库 对比报告")
print(f"时间范围: {all_dates[0]} ~ {all_dates[-1]}")
print("=" * 120)

# 总量对比
print(f"\n{'模型':<35} {'知书次数':>10} {'我方次数':>10} {'次数差':>8} {'知书$':>14} {'我方$':>14} {'差额$':>12} {'差%':>8}")
print("-" * 120)
sum_t = sum_o = 0
for m in all_models:
    t = their_total_model.get(m, {"count": 0, "quota": 0})
    o = ours_total_model.get(m, {"count": 0, "quota": 0})
    t_usd = t["quota"] / QUOTA_PER_USD
    o_usd = o["quota"] / QUOTA_PER_USD
    diff = o_usd - t_usd
    pct = diff / t_usd * 100 if t_usd else 0
    sum_t += t_usd
    sum_o += o_usd
    print(f"  {m:<33} {t['count']:>10,} {o['count']:>10,} {o['count']-t['count']:>+8,} "
          f"{t_usd:>14,.2f} {o_usd:>14,.2f} {diff:>+12,.2f} {pct:>+7.1f}%")
print("-" * 120)
total_diff = sum_o - sum_t
total_pct = total_diff / sum_t * 100 if sum_t else 0
print(f"  {'TOTAL':<33} {'':>10} {'':>10} {'':>8} "
      f"{sum_t:>14,.2f} {sum_o:>14,.2f} {total_diff:>+12,.2f} {total_pct:>+7.1f}%")

# 按日期对比
print(f"\n\n{'='*120}")
print("按日期对比（所有模型合计）")
print(f"{'='*120}")
print(f"{'日期':<10} {'知书次数':>10} {'我方次数':>10} {'次数差':>8} {'知书$':>14} {'我方$':>14} {'差额$':>12} {'差%':>8}")
print("-" * 90)
for d in all_dates:
    t_cnt = sum(v["count"] for v in their[d].values())
    o_cnt = sum(v["count"] for v in ours[d].values())
    t_usd = sum(v["quota"] for v in their[d].values()) / QUOTA_PER_USD
    o_usd = sum(v["quota"] for v in ours[d].values()) / QUOTA_PER_USD
    diff = o_usd - t_usd
    pct = diff / t_usd * 100 if t_usd else 0
    print(f"  {d:<8} {t_cnt:>10,} {o_cnt:>10,} {o_cnt-t_cnt:>+8,} "
          f"{t_usd:>14,.2f} {o_usd:>14,.2f} {diff:>+12,.2f} {pct:>+7.1f}%")

# 重点模型按日期对比
for focus_model in ["claude-opus-4-6", "claude-sonnet-4-6"]:
    print(f"\n\n{'='*120}")
    print(f"按日期对比: {focus_model}")
    print(f"{'='*120}")
    print(f"{'日期':<10} {'知书次数':>10} {'我方次数':>10} {'次数差':>8} {'知书$':>14} {'我方$':>14} {'差额$':>12} {'差%':>8}")
    print("-" * 90)
    for d in all_dates:
        t = their[d].get(focus_model, {"count": 0, "quota": 0})
        o = ours[d].get(focus_model, {"count": 0, "quota": 0})
        t_usd = t["quota"] / QUOTA_PER_USD
        o_usd = o["quota"] / QUOTA_PER_USD
        diff = o_usd - t_usd
        pct = diff / t_usd * 100 if t_usd else 0
        print(f"  {d:<8} {t['count']:>10,} {o['count']:>10,} {o['count']-t['count']:>+8,} "
              f"{t_usd:>14,.2f} {o_usd:>14,.2f} {diff:>+12,.2f} {pct:>+7.1f}%")

print(f"\n\n{'='*120}")
print("分析说明:")
print("  - 知书万卷账单中的 quota 与我方系统的 quota 均为系统计费额度")
print("  - 差异可能来自: 分段计费差异(200k prompt)、ID体系不同导致的统计口径差异")
print("  - opus-4-6 我方金额更高可能因为我方有 200k+ prompt 的高档计费")
print("  - sonnet-4-6 知书万卷金额更高可能因为他们的分段计费阈值不同")
print(f"{'='*120}")
