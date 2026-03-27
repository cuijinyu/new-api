"""临时脚本：分析知书万卷账单并与我方数据对比"""
import openpyxl
import sqlite3
import json
from collections import defaultdict
from datetime import datetime, timezone, timedelta

BILL_PATH = "reconcile/aws5101消费记录（2026-03-19 00-00-00 - 2026-03-25 00-00-00）.xlsx"
DB_PATH = "logs_analysis.db"
QUOTA_PER_USD = 500_000

BJ = timezone(timedelta(hours=8))

# ── 1. 解析知书万卷账单 ──
wb = openpyxl.load_workbook(BILL_PATH, read_only=True, data_only=True)
ws = wb[wb.sheetnames[0]]

total_rows = 0
their_model = defaultdict(lambda: {"count": 0, "quota": 0, "prompt": 0, "completion": 0})
their_user = defaultdict(lambda: {"count": 0, "quota": 0})
their_ids = set()
ts_min, ts_max = float("inf"), 0

for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0:
        continue
    total_rows += 1
    rid, user_id, created_at, typ, content, username, token_name, model_name, quota, prompt_tokens, completion_tokens = row[:11]
    other = row[-1] if len(row) >= 18 else None

    if typ != 2:
        continue

    their_ids.add(rid)
    q = quota or 0
    their_model[model_name]["count"] += 1
    their_model[model_name]["quota"] += q
    their_model[model_name]["prompt"] += (prompt_tokens or 0)
    their_model[model_name]["completion"] += (completion_tokens or 0)

    their_user[username]["count"] += 1
    their_user[username]["quota"] += q

    if created_at:
        ts_min = min(ts_min, created_at)
        ts_max = max(ts_max, created_at)

wb.close()

print(f"=== 知书万卷账单 ({BILL_PATH.split('/')[-1]}) ===")
print(f"总行数: {total_rows:,}")
dt_min = datetime.fromtimestamp(ts_min, tz=BJ)
dt_max = datetime.fromtimestamp(ts_max, tz=BJ)
print(f"时间范围: {dt_min.strftime('%Y-%m-%d %H:%M')} ~ {dt_max.strftime('%Y-%m-%d %H:%M')} (北京时间)")

print(f"\n--- 按用户 ---")
for u, s in sorted(their_user.items(), key=lambda x: -x[1]["quota"]):
    print(f"  {u}: {s['count']:,} 次, quota={s['quota']:,} (${s['quota']/QUOTA_PER_USD:,.2f})")

print(f"\n--- 按模型 ---")
total_their_quota = 0
for m, s in sorted(their_model.items(), key=lambda x: -x[1]["quota"]):
    usd = s["quota"] / QUOTA_PER_USD
    total_their_quota += s["quota"]
    print(f"  {m}: {s['count']:,} 次, ${usd:,.2f}")

total_their_usd = total_their_quota / QUOTA_PER_USD
print(f"\n总计: ${total_their_usd:,.2f}")

# ── 2. 从我方数据库查同时间段、渠道54 的数据 ──
print(f"\n\n=== 我方数据库 (channel_id=54, 同时间段) ===")
conn = sqlite3.connect(DB_PATH)

# 查所有用户在渠道54、同时间段的数据
our_data = conn.execute("""
    SELECT model_name, COUNT(*) as cnt, SUM(quota) as total_quota,
           SUM(prompt_tokens) as total_prompt, SUM(completion_tokens) as total_comp
    FROM logs
    WHERE channel_id=54 AND type=2
      AND created_at>=? AND created_at<=?
    GROUP BY model_name
    ORDER BY total_quota DESC
""", (ts_min, ts_max)).fetchall()

print(f"时间范围: {ts_min} ~ {ts_max}")
print(f"\n--- 我方全量 (ch=54, 所有用户) ---")
our_total_quota = 0
our_total_count = 0
for model, cnt, q, p, c in our_data:
    usd = q / QUOTA_PER_USD
    our_total_quota += q
    our_total_count += cnt
    print(f"  {model}: {cnt:,} 次, ${usd:,.2f}")
our_total_usd = our_total_quota / QUOTA_PER_USD
print(f"总计: {our_total_count:,} 次, ${our_total_usd:,.2f}")

# 按 user_id 分
print(f"\n--- 我方按 user_id 分 (ch=54) ---")
user_data = conn.execute("""
    SELECT user_id, username, COUNT(*) as cnt, SUM(quota) as total_quota
    FROM logs
    WHERE channel_id=54 AND type=2
      AND created_at>=? AND created_at<=?
    GROUP BY user_id
    ORDER BY total_quota DESC
""", (ts_min, ts_max)).fetchall()
for uid, uname, cnt, q in user_data:
    print(f"  user_id={uid} ({uname}): {cnt:,} 次, ${q/QUOTA_PER_USD:,.2f}")

# ── 3. 按模型对比 ──
print(f"\n\n=== 按模型对比 ===")
print(f"{'模型':<35} {'知书万卷次数':>10} {'我方次数':>10} {'差异':>8} {'知书万卷$':>12} {'我方$':>12} {'差额$':>10}")
print("-" * 110)

our_model_map = {m: (cnt, q / QUOTA_PER_USD) for m, cnt, q, p, c in our_data}
all_models = sorted(set(list(their_model.keys()) + list(our_model_map.keys())))

sum_their = sum_our = 0
for m in all_models:
    t_cnt = their_model[m]["count"] if m in their_model else 0
    t_usd = their_model[m]["quota"] / QUOTA_PER_USD if m in their_model else 0
    o_cnt, o_usd = our_model_map.get(m, (0, 0))
    diff_cnt = o_cnt - t_cnt
    diff_usd = o_usd - t_usd
    sum_their += t_usd
    sum_our += o_usd
    flag = " <!>" if abs(diff_cnt) > 10 or abs(diff_usd) > 1 else ""
    print(f"  {m:<33} {t_cnt:>10,} {o_cnt:>10,} {diff_cnt:>+8,} {t_usd:>12,.2f} {o_usd:>12,.2f} {diff_usd:>+10,.2f}{flag}")

print("-" * 110)
print(f"  {'合计':<33} {sum(s['count'] for s in their_model.values()):>10,} {our_total_count:>10,} "
      f"{our_total_count - sum(s['count'] for s in their_model.values()):>+8,} "
      f"{sum_their:>12,.2f} {sum_our:>12,.2f} {sum_our - sum_their:>+10,.2f}")

# ── 4. 检查 ID 交叉 ──
print(f"\n\n=== ID 交叉检查 ===")
our_ids_rows = conn.execute("""
    SELECT id FROM logs
    WHERE channel_id=54 AND type=2
      AND created_at>=? AND created_at<=?
""", (ts_min, ts_max)).fetchall()
our_ids = set(r[0] for r in our_ids_rows)

only_theirs = their_ids - our_ids
only_ours = our_ids - their_ids
common = their_ids & our_ids

print(f"知书万卷 ID 数: {len(their_ids):,}")
print(f"我方 ID 数: {len(our_ids):,}")
print(f"共同 ID: {len(common):,}")
print(f"仅知书万卷有: {len(only_theirs):,}")
print(f"仅我方有: {len(only_ours):,}")

if only_theirs:
    sample = list(only_theirs)[:10]
    print(f"  仅知书万卷有的 ID 样本: {sample}")

if only_ours:
    sample = list(only_ours)[:10]
    print(f"  仅我方有的 ID 样本: {sample}")
    # 查看这些 ID 的详情
    placeholders = ",".join("?" * min(len(only_ours), 20))
    sample_ids = list(only_ours)[:20]
    rows = conn.execute(f"""
        SELECT id, username, model_name, quota, created_at
        FROM logs WHERE id IN ({placeholders})
    """, sample_ids).fetchall()
    for r in rows[:10]:
        dt = datetime.fromtimestamp(r[4], tz=BJ)
        print(f"    id={r[0]} user={r[1]} model={r[2]} quota={r[3]} time={dt.strftime('%m-%d %H:%M')}")

conn.close()
