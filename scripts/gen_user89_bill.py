"""
生成用户 ID=89 的 2026年3月份账单 Excel
"""
import sqlite3
from datetime import datetime, timezone, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SQLITE_PATH = "e:/new-api/scripts/logs_analysis.db"
OUTPUT_PATH = "e:/new-api/scripts/reconcile/User89_bill_2026-03.xlsx"

USER_ID = 89
YEAR_MONTH = "2026-03"

MAR_START = int(datetime(2026, 3, 1).timestamp())
APR_START = int(datetime(2026, 4, 1).timestamp())

# 刊例价公式: quota / 500,000 = USD
QUOTA_TO_USD = 1 / 500_000

CST = timezone(timedelta(hours=8))


def fmt_ts(ts):
    if ts is None:
        return ""
    return datetime.fromtimestamp(ts, tz=CST).strftime("%Y-%m-%d %H:%M:%S")


def header_style(ws, row, cols, fill_color="1F4E79"):
    fill = PatternFill("solid", fgColor=fill_color)
    font = Font(bold=True, color="FFFFFF", size=11)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="FFFFFF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = align
        cell.border = border


def data_style(ws, row, cols, alt=False):
    fill_color = "D6E4F0" if alt else "FFFFFF"
    fill = PatternFill("solid", fgColor=fill_color)
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.border = border
        cell.alignment = align_center if col != 1 else align_left


def total_style(ws, row, cols):
    fill = PatternFill("solid", fgColor="2E75B6")
    font = Font(bold=True, color="FFFFFF", size=11)
    align = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="FFFFFF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = align
        cell.border = border


def build_summary_sheet(wb, conn):
    ws = wb.active
    ws.title = "按模型汇总"
    ws.sheet_view.showGridLines = False

    # 标题行
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = f"用户 ID={USER_ID} · {YEAR_MONTH} 月度账单汇总"
    title_cell.font = Font(bold=True, size=14, color="1F4E79")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # 生成时间
    ws.merge_cells("A2:H2")
    gen_cell = ws["A2"]
    gen_cell.value = f"生成时间: {datetime.now(CST).strftime('%Y-%m-%d %H:%M:%S CST')}"
    gen_cell.font = Font(italic=True, color="666666", size=9)
    gen_cell.alignment = Alignment(horizontal="right")

    # 表头
    headers = ["模型", "调用次数", "输入 Tokens", "输出 Tokens",
               "总 Tokens", "总 Quota", "刊例价 (USD)", "备注"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    header_style(ws, 3, len(headers))
    ws.row_dimensions[3].height = 22

    # 数据
    cur = conn.cursor()
    cur.execute("""
        SELECT model_name,
               COUNT(*) as calls,
               SUM(prompt_tokens) as pt,
               SUM(completion_tokens) as ct,
               SUM(prompt_tokens)+SUM(completion_tokens) as total_t,
               SUM(quota) as total_q
        FROM logs
        WHERE user_id=? AND created_at>=? AND created_at<?
        GROUP BY model_name
        ORDER BY total_q DESC
    """, (USER_ID, MAR_START, APR_START))
    rows = cur.fetchall()

    total_calls = total_pt = total_ct = total_tt = total_q = 0
    data_row = 4
    for i, r in enumerate(rows):
        model, calls, pt, ct, tt, q = r
        usd = (q or 0) * QUOTA_TO_USD
        ws.cell(data_row, 1, model or "(空)")
        ws.cell(data_row, 2, calls)
        ws.cell(data_row, 3, pt or 0)
        ws.cell(data_row, 4, ct or 0)
        ws.cell(data_row, 5, tt or 0)
        ws.cell(data_row, 6, q or 0)
        ws.cell(data_row, 7, round(usd, 4))
        ws.cell(data_row, 8, "")
        data_style(ws, data_row, len(headers), alt=(i % 2 == 1))
        # 数字格式
        for col in range(2, 8):
            ws.cell(data_row, col).number_format = "#,##0" if col < 7 else "#,##0.0000"
        total_calls += calls
        total_pt += pt or 0
        total_ct += ct or 0
        total_tt += tt or 0
        total_q += q or 0
        data_row += 1

    # 合计行
    total_usd = total_q * QUOTA_TO_USD
    ws.cell(data_row, 1, "合  计")
    ws.cell(data_row, 2, total_calls)
    ws.cell(data_row, 3, total_pt)
    ws.cell(data_row, 4, total_ct)
    ws.cell(data_row, 5, total_tt)
    ws.cell(data_row, 6, total_q)
    ws.cell(data_row, 7, round(total_usd, 4))
    ws.cell(data_row, 8, "")
    total_style(ws, data_row, len(headers))
    for col in range(2, 8):
        ws.cell(data_row, col).number_format = "#,##0" if col < 7 else "#,##0.0000"
    ws.row_dimensions[data_row].height = 22

    # 列宽
    col_widths = [42, 12, 16, 16, 16, 18, 16, 12]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    return total_calls, total_pt, total_ct, total_tt, total_q, total_usd


def build_daily_sheet(wb, conn):
    ws = wb.create_sheet("按日汇总")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = f"用户 ID={USER_ID} · {YEAR_MONTH} 按日汇总"
    title_cell.font = Font(bold=True, size=14, color="1F4E79")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    headers = ["日期", "调用次数", "输入 Tokens", "输出 Tokens", "总 Tokens", "总 Quota", "费用 (USD)"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=h)
    header_style(ws, 2, len(headers))
    ws.row_dimensions[2].height = 22

    cur = conn.cursor()
    # 按天分组（UTC+8）
    cur.execute("""
        SELECT created_at, COUNT(*), SUM(prompt_tokens), SUM(completion_tokens), SUM(quota)
        FROM logs
        WHERE user_id=? AND created_at>=? AND created_at<?
        GROUP BY (created_at + 28800) / 86400
        ORDER BY created_at
    """, (USER_ID, MAR_START, APR_START))
    rows = cur.fetchall()

    data_row = 3
    for i, r in enumerate(rows):
        ts, calls, pt, ct, q = r
        day_str = datetime.fromtimestamp(ts, tz=CST).strftime("%Y-%m-%d")
        tt = (pt or 0) + (ct or 0)
        usd = (q or 0) * QUOTA_TO_USD
        ws.cell(data_row, 1, day_str)
        ws.cell(data_row, 2, calls)
        ws.cell(data_row, 3, pt or 0)
        ws.cell(data_row, 4, ct or 0)
        ws.cell(data_row, 5, tt)
        ws.cell(data_row, 6, q or 0)
        ws.cell(data_row, 7, round(usd, 4))
        data_style(ws, data_row, len(headers), alt=(i % 2 == 1))
        for col in range(2, 8):
            ws.cell(data_row, col).number_format = "#,##0" if col < 7 else "#,##0.0000"
        data_row += 1

    col_widths = [14, 12, 16, 16, 16, 18, 14]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w


def build_detail_sheet(wb, conn):
    ws = wb.create_sheet("调用明细")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:K1")
    title_cell = ws["A1"]
    title_cell.value = f"用户 ID={USER_ID} · {YEAR_MONTH} 调用明细（最多导出 50,000 条）"
    title_cell.font = Font(bold=True, size=13, color="1F4E79")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    headers = ["时间 (CST)", "模型", "调用类型", "输入 Tokens", "输出 Tokens",
               "总 Tokens", "Quota", "费用 (USD)", "渠道", "Token名称", "IP"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=h)
    header_style(ws, 2, len(headers))
    ws.row_dimensions[2].height = 22

    cur = conn.cursor()
    cur.execute("""
        SELECT created_at, model_name, type,
               prompt_tokens, completion_tokens,
               quota, channel_name, token_name, ip
        FROM logs
        WHERE user_id=? AND created_at>=? AND created_at<?
        ORDER BY created_at DESC
        LIMIT 50000
    """, (USER_ID, MAR_START, APR_START))
    rows = cur.fetchall()

    type_map = {1: "充值", 2: "消费", 3: "管理", 4: "系统"}

    data_row = 3
    for i, r in enumerate(rows):
        ts, model, typ, pt, ct, q, ch_name, tok_name, ip = r
        tt = (pt or 0) + (ct or 0)
        usd = (q or 0) * QUOTA_TO_USD
        ws.cell(data_row, 1, fmt_ts(ts))
        ws.cell(data_row, 2, model or "")
        ws.cell(data_row, 3, type_map.get(typ, str(typ)))
        ws.cell(data_row, 4, pt or 0)
        ws.cell(data_row, 5, ct or 0)
        ws.cell(data_row, 6, tt)
        ws.cell(data_row, 7, q or 0)
        ws.cell(data_row, 8, round(usd, 6))
        ws.cell(data_row, 9, ch_name or "")
        ws.cell(data_row, 10, tok_name or "")
        ws.cell(data_row, 11, ip or "")
        data_style(ws, data_row, len(headers), alt=(i % 2 == 1))
        for col in [4, 5, 6, 7]:
            ws.cell(data_row, col).number_format = "#,##0"
        ws.cell(data_row, 8).number_format = "#,##0.000000"
        data_row += 1

    col_widths = [20, 36, 8, 14, 14, 14, 14, 14, 20, 20, 16]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = "A3"


def main():
    conn = sqlite3.connect(SQLITE_PATH)

    wb = openpyxl.Workbook()
    print("[*] 生成按模型汇总 ...")
    total_calls, total_pt, total_ct, total_tt, total_q, total_usd = build_summary_sheet(wb, conn)

    print("[*] 生成按日汇总 ...")
    build_daily_sheet(wb, conn)

    print("[*] 生成调用明细 ...")
    build_detail_sheet(wb, conn)

    conn.close()

    wb.save(OUTPUT_PATH)
    print(f"\n[完成] 账单已保存: {OUTPUT_PATH}")
    print(f"  总调用次数: {total_calls:,}")
    print(f"  总输入 Tokens: {total_pt:,}")
    print(f"  总输出 Tokens: {total_ct:,}")
    print(f"  总 Tokens: {total_tt:,}")
    print(f"  总 Quota: {total_q:,}")
    print(f"  刊例价 (USD): {total_usd:,.4f}")


if __name__ == "__main__":
    main()
