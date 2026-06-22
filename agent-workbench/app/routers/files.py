"""文件上传、下载与预览路由。"""

from __future__ import annotations

import base64
import binascii
import csv
import hashlib
import json
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any

import psycopg2.extras
from fastapi import APIRouter, HTTPException, Query
from openpyxl import load_workbook
from pydantic import BaseModel, Field

from .. import security
from ..services.artifacts import artifacts, billing_artifacts, stream_download_response
from ..services.core import (
    clamp_int,
    db_conn,
    ensure_agent_session,
    fetch_all,
    fetch_one,
    insert_agent_event,
    json_safe,
    new_id,
    safe_filename,
    slug,
    utc_now,
)

router = APIRouter(tags=["files"])

MAX_PREVIEW_BYTES = 2 * 1024 * 1024
MAX_SHEET_ROWS = 200
MAX_SHEET_COLS = 50
TEXT_EXTENSIONS = {".md", ".txt", ".log", ".yaml", ".yml", ".csv", ".tsv", ".jsonl"}
JSON_EXTENSIONS = {".json"}
XLSX_EXTENSIONS = {".xlsx", ".xls"}
IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".gif", ".webp", ".svg"}
PDF_EXTENSIONS = {".pdf"}


class FileUploadJsonRequest(BaseModel):
    filename: str
    content_base64: str
    content_type: str = "application/octet-stream"
    category: str = "general"
    job_id: str | None = None
    session_id: str | None = None
    uploaded_by: str = "user"
    metadata: dict[str, Any] = Field(default_factory=dict)


def _resolve_store(s3_uri: str):
    if billing_artifacts is not artifacts and s3_uri.startswith(f"s3://{billing_artifacts.bucket}/"):
        return billing_artifacts
    return artifacts


def _load_file_row(file_id: str) -> dict[str, Any]:
    with db_conn() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            row = fetch_one(cur, "SELECT * FROM uploaded_files WHERE id = %s", (file_id,))
            if not row:
                raise HTTPException(status_code=404, detail="file not found")
            return dict(row)


def _file_extension(filename: str, content_type: str | None = None) -> str:
    ext = Path(filename or "").suffix.lower()
    if ext:
        return ext
    ctype = (content_type or "").lower()
    if "pdf" in ctype:
        return ".pdf"
    if "json" in ctype:
        return ".json"
    if "csv" in ctype:
        return ".csv"
    if "spreadsheet" in ctype or "excel" in ctype:
        return ".xlsx"
    if ctype.startswith("image/"):
        return ".png"
    if ctype.startswith("text/"):
        return ".txt"
    return ""


def _access_url(row: dict[str, Any], store, file_id: str, *, inline: bool = False) -> str:
    # 浏览器侧统一走 API 代理，避免 presigned 指向内网 S3 endpoint。
    disposition = "inline" if inline else "attachment"
    return f"/api/files/{file_id}/download?disposition={disposition}"


def _read_file_bytes(row: dict[str, Any], store, max_bytes: int) -> bytes:
    target_uri = str(row.get("s3_uri") or "")
    if not target_uri:
        raise HTTPException(status_code=400, detail="file has no storage uri")
    raw = store.get_bytes(target_uri)
    if len(raw) > max_bytes:
        return raw[:max_bytes]
    return raw


def _preview_csv(text: str, truncated: bool) -> dict[str, Any]:
    reader = csv.reader(StringIO(text))
    rows = list(reader)
    columns = rows[0] if rows else []
    data_rows = rows[1:] if len(rows) > 1 else []
    if len(data_rows) > MAX_SHEET_ROWS:
        data_rows = data_rows[:MAX_SHEET_ROWS]
        truncated = True
    if columns and len(columns) > MAX_SHEET_COLS:
        columns = columns[:MAX_SHEET_COLS]
        data_rows = [row[:MAX_SHEET_COLS] for row in data_rows]
        truncated = True
    return {"kind": "csv", "columns": columns, "rows": data_rows, "truncated": truncated}


def _preview_xlsx(raw: bytes) -> dict[str, Any]:
    workbook = load_workbook(BytesIO(raw), read_only=True, data_only=True)
    sheets: list[dict[str, Any]] = []
    try:
        for worksheet in workbook.worksheets:
            rows: list[list[str]] = []
            truncated = False
            for index, row in enumerate(worksheet.iter_rows(values_only=True)):
                if index >= MAX_SHEET_ROWS:
                    truncated = True
                    break
                cells = ["" if value is None else str(value) for value in row[:MAX_SHEET_COLS]]
                if len(row) > MAX_SHEET_COLS:
                    truncated = True
                rows.append(cells)
            columns = [chr(65 + index) if index < 26 else f"C{index + 1}" for index in range(len(rows[0]) if rows else 0)]
            sheets.append({"name": worksheet.title, "columns": columns, "rows": rows, "truncated": truncated})
    finally:
        workbook.close()
    return {"kind": "sheet", "sheets": sheets}


def _build_preview(row: dict[str, Any], store) -> dict[str, Any]:
    file_id = str(row.get("id") or "")
    filename = str(row.get("filename") or "download")
    content_type = str(row.get("content_type") or "application/octet-stream")
    byte_size = int(row.get("byte_size") or 0)
    ext = _file_extension(filename, content_type)

    base = {
        "file_id": file_id,
        "filename": filename,
        "content_type": content_type,
        "byte_size": byte_size,
        "s3_uri": row.get("s3_uri"),
    }

    if ext in PDF_EXTENSIONS or content_type == "application/pdf":
        return {**base, "kind": "pdf", "url": _access_url(row, store, file_id, inline=True)}
    if ext in IMAGE_EXTENSIONS or content_type.startswith("image/"):
        return {**base, "kind": "image", "url": _access_url(row, store, file_id, inline=True)}

    if ext in XLSX_EXTENSIONS or "spreadsheet" in content_type or "excel" in content_type:
        raw = _read_file_bytes(row, store, MAX_PREVIEW_BYTES)
        truncated = byte_size > len(raw)
        payload = _preview_xlsx(raw)
        return {**base, **payload, "truncated": truncated or any(sheet.get("truncated") for sheet in payload.get("sheets", []))}

    raw = _read_file_bytes(row, store, MAX_PREVIEW_BYTES)
    truncated = byte_size > len(raw)
    text = raw.decode("utf-8", errors="replace")

    if ext in JSON_EXTENSIONS or content_type == "application/json":
        try:
            parsed = json.loads(text)
            return {**base, "kind": "json", "data": json_safe(parsed), "truncated": truncated}
        except json.JSONDecodeError:
            pass

    if ext in {".csv", ".tsv"} or content_type in {"text/csv", "application/csv"}:
        return {**base, **_preview_csv(text, truncated)}

    if ext in TEXT_EXTENSIONS or content_type.startswith("text/"):
        return {**base, "kind": "text", "text": text, "truncated": truncated}

    return {
        **base,
        "kind": "binary",
        "download_url": _access_url(row, store, file_id, inline=False),
        "message": "此文件类型不支持在线预览，请下载后查看",
    }


@router.post("/api/files/upload")
def upload_file_json(req: FileUploadJsonRequest) -> dict[str, Any]:
    filename = safe_filename(req.filename)
    if not filename:
        raise HTTPException(status_code=400, detail="filename is required")
    try:
        raw = base64.b64decode(req.content_base64, validate=True)
    except (binascii.Error, ValueError) as exc:
        raise HTTPException(status_code=400, detail="content_base64 is not valid base64") from exc
    if not raw:
        raise HTTPException(status_code=400, detail="file is empty")
    if len(raw) > security.max_upload_bytes():
        raise HTTPException(status_code=413, detail=f"file exceeds max upload size {security.max_upload_bytes()} bytes")

    file_id = new_id("file")
    digest = hashlib.sha256(raw).hexdigest()
    category = slug(req.category or "general")
    prefix = f"uploads/{utc_now().date().isoformat()}/{category}/{file_id}"
    s3_uri = artifacts.put_bytes(f"{prefix}/{filename}", raw, req.content_type or "application/octet-stream")

    with db_conn() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            if req.job_id:
                job = fetch_one(cur, "SELECT id FROM jobs WHERE id = %s", (req.job_id,))
                if not job:
                    raise HTTPException(status_code=404, detail="job not found")
            if req.session_id:
                ensure_agent_session(cur, req.session_id)
            cur.execute(
                """
                INSERT INTO uploaded_files (
                    id, filename, content_type, byte_size, sha256, category,
                    job_id, session_id, s3_uri, uploaded_by, metadata
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING *
                """,
                (file_id, filename, req.content_type or "application/octet-stream", len(raw), digest, category, req.job_id, req.session_id, s3_uri, req.uploaded_by, psycopg2.extras.Json(req.metadata)),
            )
            row = dict(cur.fetchone())
            if req.session_id:
                insert_agent_event(
                    cur, req.session_id, "message", "user",
                    f"已上传对账资料：{filename}，类型：{category}，地址：{s3_uri}",
                    {"source": "file_upload", "file_id": file_id, "filename": filename, "category": category, "s3_uri": s3_uri, **req.metadata},
                )
                cur.execute("UPDATE agent_sessions SET updated_at = NOW() WHERE id = %s", (req.session_id,))

    return {"status": "uploaded", "file": row}


@router.get("/api/files")
def list_uploaded_files(job_id: str | None = None, session_id: str | None = None) -> dict[str, Any]:
    filters: list[str] = []
    args: list[Any] = []
    if job_id:
        filters.append("job_id = %s")
        args.append(job_id)
    if session_id:
        filters.append("session_id = %s")
        args.append(session_id)
    where = f"WHERE {' AND '.join(filters)}" if filters else ""
    with db_conn() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            rows = fetch_all(cur, f"SELECT * FROM uploaded_files {where} ORDER BY created_at DESC LIMIT 100", tuple(args))
    return {"items": rows}


@router.get("/api/files/{file_id}/download")
def download_file(file_id: str, disposition: str = Query(default="attachment")):
    row = _load_file_row(file_id)
    target_uri = str(row.get("s3_uri") or "")
    if not target_uri:
        raise HTTPException(status_code=400, detail="file has no storage uri")
    filename = str(row.get("filename") or "download")
    content_type = str(row.get("content_type") or "application/octet-stream")
    store = _resolve_store(target_uri)
    if not filename or filename == "download":
        filename = store.download_filename(target_uri, filename)
    inline = disposition.strip().lower() == "inline"
    return stream_download_response(store, target_uri, filename, inline=inline, content_type=content_type if inline else None)


@router.get("/api/files/{file_id}/preview")
def preview_file(file_id: str, max_bytes: int = Query(default=MAX_PREVIEW_BYTES)) -> dict[str, Any]:
    row = _load_file_row(file_id)
    target_uri = str(row.get("s3_uri") or "")
    if not target_uri:
        raise HTTPException(status_code=400, detail="file has no storage uri")
    clamp_int(max_bytes, MAX_PREVIEW_BYTES, 4 * 1024, MAX_PREVIEW_BYTES)
    store = _resolve_store(target_uri)
    return _build_preview(row, store)
